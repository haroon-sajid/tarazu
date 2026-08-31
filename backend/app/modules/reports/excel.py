"""Render a `ReportContent` to an Excel workbook with openpyxl.

One sheet for the summary and one per section, with the same rows the PDF
carries. Cells are written as text exactly as the content builder produced
them, so nothing in the workbook is a formula and nothing can be recomputed
by opening it.
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.reports.content import ReportContent, TableSection

__all__ = ["render_excel"]

_HEAD_FILL = PatternFill("solid", fgColor="0E7C66")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_LABEL_FONT = Font(bold=True, color="6B7A8A")
_WRAP = Alignment(wrap_text=True, vertical="top")

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sheet_title(title: str, used: set[str]) -> str:
    base = _INVALID_SHEET_CHARS.sub(" ", title).strip()[:31] or "Sheet"
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f" ({counter})"
        candidate = base[: 31 - len(suffix)] + suffix
        counter += 1
    used.add(candidate)
    return candidate


def _write_section(workbook: Workbook, section: TableSection, used: set[str]) -> None:
    sheet = workbook.create_sheet(_sheet_title(section.title, used))
    row = 1
    sheet.cell(row=row, column=1, value=section.title).font = Font(bold=True, size=13)
    row += 1
    if section.note:
        sheet.cell(row=row, column=1, value=section.note).alignment = _WRAP
        row += 1
    row += 1

    for index, column in enumerate(section.columns, start=1):
        cell = sheet.cell(row=row, column=index, value=column)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = _WRAP
    header_row = row
    row += 1

    for values in section.rows:
        for index, value in enumerate(values, start=1):
            sheet.cell(row=row, column=index, value=value).alignment = _WRAP
        row += 1
    if not section.rows:
        sheet.cell(row=row, column=1, value="Nothing to report.")

    weights = section.widths or [1.0] * len(section.columns)
    for index, weight in enumerate(weights, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(12, min(70, int(16 * weight)))
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


def _restamp(archive: bytes, stamp: datetime) -> bytes:
    """Rewrite the zip with every entry dated `stamp`, so the bytes are reproducible.

    openpyxl dates each zip entry with the wall clock as it writes, which
    would make two renderings of the same report differ by a few bytes and
    give the report record two digests for one document.
    """
    date_time = (max(stamp.year, 1980), stamp.month, stamp.day, stamp.hour, stamp.minute,
                 stamp.second - stamp.second % 2)
    iso = stamp.strftime("%Y-%m-%dT%H:%M:%SZ")
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(archive)) as source, zipfile.ZipFile(
        out, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for entry in source.infolist():
            data = source.read(entry.filename)
            if entry.filename == "docProps/core.xml":
                # openpyxl writes the wall clock here whatever the properties
                # say; the report's own time is the only one that belongs.
                text = data.decode("utf-8")
                for tag in ("dcterms:created", "dcterms:modified"):
                    text = re.sub(
                        rf"(<{tag}[^>]*>)[^<]*(</{tag}>)", rf"\g<1>{iso}\g<2>", text
                    )
                data = text.encode("utf-8")
            info = zipfile.ZipInfo(entry.filename, date_time=date_time)
            info.compress_type = zipfile.ZIP_DEFLATED
            target.writestr(info, data)
    return out.getvalue()


def render_excel(content: ReportContent) -> bytes:
    """The whole report as `.xlsx` bytes.

    Byte-for-byte reproducible from the same content: the workbook's own
    created/modified stamps and every zip entry's date are set to the report's
    generation time rather than the wall clock, so the digest on the report
    record is a property of the report, not of the second it was rendered in.
    """
    workbook = Workbook()
    generated_at = content.meta.generated_at.replace(tzinfo=None)
    workbook.properties.creator = "Tarazu — AI Audit Assistant"
    workbook.properties.created = generated_at
    workbook.properties.modified = generated_at
    workbook.properties.title = f"Tarazu report {content.meta.report_id}"
    used: set[str] = set()

    summary = workbook.active
    summary.title = _sheet_title("Summary", used)
    summary.cell(row=1, column=1, value="Tarazu — Audit Reconciliation Report").font = Font(
        bold=True, size=14
    )
    summary.cell(
        row=2,
        column=1,
        value=(
            f"{content.meta.client_name} · {content.meta.case_id} · "
            f"{content.meta.report_id}"
        ),
    ).font = Font(color="6B7A8A")
    row = 4
    for label, value in content.summary:
        summary.cell(row=row, column=1, value=label).font = _LABEL_FONT
        summary.cell(row=row, column=2, value=value).alignment = _WRAP
        row += 1
    row += 1
    summary.cell(row=row, column=1, value=content.closing).alignment = _WRAP
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 100

    for section in content.sections:
        _write_section(workbook, section, used)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return _restamp(buffer.getvalue(), generated_at)
