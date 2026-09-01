"""Public interface of the analytics module.

This is the only file other modules may import from `modules/analytics/`. It
accepts and returns `app/shared/` schema objects exclusively — never raw dicts.

What this module does:

- Reads a sales export in whatever shape the client's software produced —
  CSV, TSV or another delimited text file, an Excel workbook (`.xlsx`, `.xlsm`,
  `.xls`), an OpenDocument spreadsheet, or JSON — into `SalesRecord` objects.
  **No AI on this path**: a spreadsheet is already structured, and a model
  could only misread a number that is sitting right there in a cell. This is
  the same no-AI path the ledger takes.
- Cleans it the way an auditor would by hand, and says what it did: finds the
  header under any title rows, picks the worksheet that holds the table, maps
  the client's own column names, decodes non-UTF-8 files, derives an amount
  from quantity × unit price when no amount column exists, skips blank and
  total rows, and files rows with no customer or product under "Unspecified"
  instead of dropping them. Every one of those decisions is reported in a
  `SourceReadReport`, never taken silently.
- Aggregates the records into a `SalesAnalyticsResult`: revenue by month, by
  product, by region, the top customers, and anomalies worth a human's
  attention. Every figure is a sum or a count over the rows that were read;
  money stays `Decimal` end to end, so the breakdowns add back to the total
  exactly.
- Renders a saved readout as an Excel workbook (`export_workbook`) so the
  insights can leave the product as a file.

What it must never do (see the module README and CLAUDE.md): import an AI
client, call anything over the network, or turn an anomaly into a verdict. The
anomalies are rule findings — `negative-amount`, `duplicate-transaction`,
`revenue-spike`, `large-transaction` — and each names the row or month it is
about, the same way `rules/` flags name theirs.

Typical use::

    records, report = read_sales_export("SLS-0001", "sales.xlsx", content)
    result = analyze_sales(records, reports=[report])
    workbook_bytes = export_workbook(result)
"""

from __future__ import annotations

import csv
import io
import json
import logging
import re
import statistics
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.shared.schemas import (
    Anomaly,
    CustomerSummary,
    MonthlyRevenue,
    ProductRevenue,
    Provenance,
    RegionSummary,
    SalesAnalyticsResult,
    SalesRecord,
    SourceReadReport,
)

__all__ = [
    "ANOMALY_KINDS",
    "LARGE_TRANSACTION_FACTOR",
    "LARGE_TRANSACTION_MIN_SAMPLE",
    "SPIKE_FACTOR",
    "SPIKE_MIN_MONTHS",
    "SUPPORTED_SUFFIXES",
    "TOP_CUSTOMERS",
    "UNSPECIFIED",
    "SalesReadError",
    "analyze_sales",
    "export_workbook",
    "read_sales_data",
    "read_sales_export",
]

logger = logging.getLogger(__name__)


class SalesReadError(ValueError):
    """The sales data could not be read: an unsupported format, no header
    naming a date and an amount, a malformed file, or no usable rows."""


#: The file types the reader understands. The API refuses anything else before
#: it is stored, and the upload screen offers the same list.
SUPPORTED_SUFFIXES = frozenset(
    {".csv", ".tsv", ".txt", ".xlsx", ".xlsm", ".xls", ".ods", ".json"}
)

#: The label a row is filed under when the export has no customer or product
#: column, or the cell is empty. The row still counts toward revenue; it is not
#: dropped, and the report says how many rows were filed this way.
UNSPECIFIED = "Unspecified"

#: How many customers the readout ranks. The schema caps the list at the same
#: number, so the contract and the computation cannot drift apart.
TOP_CUSTOMERS = 5

#: A month is only called a spike against at least this many months of
#: context. Two months cannot make a trend, and one of two is always "above
#: the median".
SPIKE_MIN_MONTHS = 3

#: A month whose revenue is more than this multiple of the median month — or
#: less than the median divided by it — is flagged as a spike.
SPIKE_FACTOR = Decimal("2")

#: A whale transaction is only looked for once there is this much of a sample,
#: for the same reason Benford refuses significance below fifty rows.
LARGE_TRANSACTION_MIN_SAMPLE = 10

#: A sale more than this multiple of the median sale is flagged. A multiple of
#: the median rather than of the mean, so one enormous sale cannot raise the
#: bar against its own detection.
LARGE_TRANSACTION_FACTOR = Decimal("10")

#: The anomaly kinds this module emits, in the order they are evaluated. A
#: plain string rather than an enum, matching `rule_id` in `rules/`.
ANOMALY_KINDS = (
    "negative-amount",
    "duplicate-transaction",
    "revenue-spike",
    "large-transaction",
)


# --------------------------------------------------------------------------- #
# 1. Reading the export
# --------------------------------------------------------------------------- #
#
# The mirror of `extraction/ledger_reader.py`, which this module deliberately
# does not import: a module's internals are its own, so the parsing helpers
# live here too.


#: Header aliases seen in real sales exports (ERPs, POS dumps, Excel by hand).
#: Compared after normalising to lowercase with underscores. An exact alias
#: match is tried first; the looser `_HINTS` below only fill in what is left.
_ALIASES: dict[str, tuple[str, ...]] = {
    "date": (
        "date", "sale_date", "sales_date", "txn_date", "transaction_date",
        "order_date", "invoice_date", "dated", "date_of_sale", "posting_date",
        "trans_date", "bill_date", "doc_date", "document_date", "voucher_date",
        "period", "day",
    ),
    "amount": (
        "amount", "amt", "value", "revenue", "total", "sale_amount",
        "amount_pkr", "line_total", "net_amount", "gross_amount", "sales",
        "sale_value", "sales_value", "grand_total", "total_amount", "net_total",
        "invoice_amount", "invoice_total", "bill_amount", "net", "subtotal",
        "sub_total", "amount_rs", "total_pkr", "sales_amount", "turnover",
    ),
    "quantity": (
        "quantity", "qty", "units", "pcs", "pieces", "nos", "no_of_units",
        "qnty", "quantity_sold", "qty_sold",
    ),
    "unit_price": (
        "unit_price", "price", "rate", "unit_rate", "price_per_unit",
        "selling_price", "sale_price", "rate_pkr", "price_pkr", "unit_sale_price",
    ),
    "customer_name": (
        "customer_name", "customer", "client", "buyer", "party_name", "party",
        "account_name", "name", "cust", "cust_name", "client_name",
        "customer_id", "buyer_name", "account", "dealer", "distributor",
        "retailer", "sold_to", "bill_to",
    ),
    "product": (
        "product", "product_name", "item", "item_name", "sku", "description",
        "particulars", "item_description", "product_code", "category", "goods",
        "article", "design", "material", "product_description", "item_code",
    ),
    "region": (
        "region", "city", "area", "territory", "zone", "province", "location",
        "branch", "market", "state", "country", "district", "city_name",
        "store", "outlet", "warehouse",
    ),
    "sales_row_id": (
        "sales_row_id", "id", "ref", "reference", "invoice_no", "invoice_number",
        "order_no", "entry_no", "invoice", "order_id", "transaction_id",
        "txn_id", "receipt_no", "bill_no", "voucher_no", "sr_no", "s_no", "sno",
        "serial", "row_id", "doc_no",
    ),
}

#: Substrings that identify a column when no alias matched exactly —
#: `Sale Date (DD/MM)`, `Net Sales Amount`, `Customer / Party`. Only unresolved
#: fields are filled this way, first matching column wins.
_HINTS: dict[str, tuple[str, ...]] = {
    "date": ("date",),
    "amount": ("amount", "total", "revenue", "value", "sales"),
    "quantity": ("qty", "quantity"),
    "unit_price": ("price", "rate"),
    "customer_name": ("customer", "client", "buyer", "party"),
    "product": ("product", "item", "sku", "description"),
    "region": ("region", "city", "province", "territory", "branch", "area", "zone"),
}

#: A money-looking header that is not the sale itself. A hint never lands on
#: one of these, so `Total Tax` or `Discount Amount` cannot become the revenue.
_NOT_MONEY = ("tax", "vat", "gst", "discount", "cost", "balance", "due",
              "outstanding", "paid", "received", "commission", "margin")

#: How many leading rows are searched for the header. Title blocks on real
#: exports run to a handful of lines; thirty is generous.
_HEADER_SCAN_ROWS = 30

#: The delimiters a text export may be split on, in the order the sniffer
#: prefers them.
_DELIMITERS = ",;\t|"

#: Total and subtotal lines, which exports append and which must never be
#: counted as a sale.
_TOTAL_ROW = re.compile(r"^(grand\s*)?(sub\s*)?totals?\b", re.IGNORECASE)

#: The numeric core of a money cell, anchoring on the digits so `Rs. 45,900/-`
#: survives the same way it does in the ledger reader.
_MONEY_NUMBER = re.compile(r"\d[\d,\s]*(?:\.\d+)?")

#: An ISO date (`2026-06-02`, with or without a time). Unambiguous, so the
#: day-first convention must not be applied to it — pandas otherwise reads
#: the month as the day.
_ISO_DATE = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}")

#: Excel stores a date as days since 1899-12-30. Serials in this band are the
#: years 1954–2119; anything else is not a date that was lost its format.
_EXCEL_EPOCH = datetime(1899, 12, 30)
_SERIAL_MIN, _SERIAL_MAX = 20000, 80000


@dataclass
class _Table:
    """One grid of cells read from a file, header not yet located."""

    rows: list[list[object]]
    format: str
    sheet: str | None = None
    encoding: str | None = None
    delimiter: str | None = None
    #: Width of each row as it came off the file, before padding — the ragged
    #: check needs the original count.
    raw_widths: list[int] = field(default_factory=list)


def _suffix(filename: str) -> str:
    return "." + filename.rsplit(".", 1)[1].lower() if "." in filename else ""


def _normalise_header(name: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(name).strip().lower()).strip("_")


def _text(value: object) -> str:
    """A cell as text: empty for missing, otherwise stripped with inner
    whitespace collapsed — `Gulberg  Traders ` and `Gulberg Traders` are the
    same customer, and grouping on them separately would split one party's
    revenue in two."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return " ".join(str(value).split())


def _cell(value: object) -> object:
    """Normalise a raw spreadsheet cell: NaN becomes empty, text is trimmed,
    numbers and datetimes pass through untouched for the typed parsers."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _decode_text(content: bytes) -> tuple[str, str]:
    """Decode a text export, trying the encodings real exports arrive in.

    UTF-8 (with or without Excel's BOM) first; UTF-16 only when its byte-order
    mark says so, because a UTF-16 decode of anything else "succeeds" as
    garbage; then Windows-1252, which is what Excel's "CSV" save produces on a
    Pakistani desktop; and Latin-1 last, which cannot fail.
    """
    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        return content.decode("utf-16"), "utf-16"
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise SalesReadError("the sales data could not be decoded as text")


def _sniff_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters=_DELIMITERS).delimiter
    except csv.Error:
        counts = {delimiter: sample.count(delimiter) for delimiter in _DELIMITERS}
        best = max(counts, key=lambda key: counts[key])
        return best if counts[best] > 0 else ","


def _load_delimited(content: bytes, filename: str, suffix: str) -> list[_Table]:
    text, encoding = _decode_text(content)
    delimiter = "\t" if suffix == ".tsv" else _sniff_delimiter(text)
    rows = list(csv.reader(io.StringIO(text, newline=""), delimiter=delimiter))
    widths = [len(row) for row in rows]
    width = max(widths, default=0)
    padded: list[list[object]] = [
        [cell.strip() for cell in row] + [""] * (width - len(row)) for row in rows
    ]
    return [
        _Table(
            rows=padded,
            format="tsv" if delimiter == "\t" else "csv",
            encoding=encoding,
            delimiter=delimiter,
            raw_widths=widths,
        )
    ]


def _load_workbook(content: bytes, filename: str, suffix: str) -> list[_Table]:
    engine = "odf" if suffix == ".ods" else None
    try:
        sheets = pd.read_excel(
            io.BytesIO(content), sheet_name=None, header=None, dtype=object, engine=engine
        )
    except ImportError as exc:
        raise SalesReadError(
            f"reading {filename!r} needs a spreadsheet library that is not "
            f"installed on this server: {exc}"
        ) from exc
    except Exception as exc:  # noqa: BLE001 - corrupt, encrypted, or not a workbook at all
        raise SalesReadError(
            f"{filename!r} could not be opened as a spreadsheet: {exc}"
        ) from exc

    tables: list[_Table] = []
    for name, frame in sheets.items():
        rows = [
            [_cell(value) for value in row]
            for row in frame.itertuples(index=False, name=None)
        ]
        tables.append(
            _Table(
                rows=rows,
                format="ods" if engine == "odf" else "excel",
                sheet=str(name),
                raw_widths=[len(row) for row in rows],
            )
        )
    return tables


def _json_records(payload: object) -> list[dict]:
    """The list of row objects inside a JSON export, wherever it was put."""
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict):
        for key in ("data", "rows", "records", "items", "sales", "results", "values"):
            value = payload.get(key)
            if isinstance(value, list):
                return [row for row in value if isinstance(row, dict)]
        for value in payload.values():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                return [row for row in value if isinstance(row, dict)]
    return []


def _load_json(content: bytes, filename: str) -> list[_Table]:
    text, encoding = _decode_text(content)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise SalesReadError(f"{filename!r} is not valid JSON: {exc}") from exc
    records = _json_records(payload)
    if not records:
        raise SalesReadError(
            f"{filename!r} holds no sales records: expected a JSON list of "
            "objects, or an object with one under a key such as \"data\""
        )
    headers: list[str] = []
    for record in records:
        for key in record:
            if key not in headers:
                headers.append(str(key))
    rows: list[list[object]] = [list(headers)]
    rows.extend([_cell(record.get(header)) for header in headers] for record in records)
    return [
        _Table(rows=rows, format="json", encoding=encoding, raw_widths=[len(headers)] * len(rows))
    ]


def _load_tables(content: bytes, filename: str, suffix: str) -> list[_Table]:
    if suffix in (".csv", ".tsv", ".txt"):
        return _load_delimited(content, filename, suffix)
    if suffix in (".xlsx", ".xlsm", ".xls", ".ods"):
        return _load_workbook(content, filename, suffix)
    if suffix == ".json":
        return _load_json(content, filename)
    raise SalesReadError(  # pragma: no cover - guarded by SUPPORTED_SUFFIXES
        f"unsupported sales data format: {filename!r}"
    )


def _resolve_headers(headers: list[str]) -> dict[str, int]:
    """Map canonical field names onto column positions of one candidate header
    row: exact aliases first, then substring hints for whatever is left."""
    normalised = [_normalise_header(header) for header in headers]
    resolved: dict[str, int] = {}
    taken: set[int] = set()

    for canonical, aliases in _ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                position = normalised.index(alias)
                if position not in taken:
                    resolved[canonical] = position
                    taken.add(position)
                    break

    for canonical, hints in _HINTS.items():
        if canonical in resolved:
            continue
        money_like = canonical in ("amount", "unit_price")
        for position, header in enumerate(normalised):
            if position in taken or not header:
                continue
            if money_like and any(word in header for word in _NOT_MONEY):
                continue
            if any(hint in header for hint in hints):
                resolved[canonical] = position
                taken.add(position)
                break
    return resolved


def _has_required(mapping: dict[str, int]) -> bool:
    """A table needs a date and a way to get an amount; everything else has a
    sensible fallback."""
    return "date" in mapping and (
        "amount" in mapping or ("quantity" in mapping and "unit_price" in mapping)
    )


def _find_header(rows: list[list[object]]) -> tuple[int, dict[str, int]] | None:
    """The header row and its column map: the row among the first thirty that
    names the most known fields, provided it names a date and an amount."""
    best: tuple[int, int, dict[str, int]] | None = None
    for index, row in enumerate(rows[:_HEADER_SCAN_ROWS]):
        headers = [_text(cell) for cell in row]
        if sum(1 for header in headers if header) < 2:
            continue
        mapping = _resolve_headers(headers)
        if not _has_required(mapping):
            continue
        score = len(mapping)
        if best is None or score > best[0]:
            best = (score, index, mapping)
    return None if best is None else (best[1], best[2])


def _to_decimal(raw: object) -> Decimal | None:
    """Parse a money cell. Handles 'Rs. 49,500/-', '(1,200)', and plain numbers."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float, Decimal)):
        try:
            return Decimal(str(raw))
        except InvalidOperation:
            return None

    text = str(raw).strip()
    if not text:
        return None

    match = _MONEY_NUMBER.search(text)
    if not match:
        return None

    digits = match.group(0).replace(",", "").replace(" ", "").rstrip(".")
    if not digits:
        return None

    # Accounting style writes negatives as `(1,200)`; a bare `-` also counts,
    # but only when it sits before the digits. A trailing `/-` is the Pakistani
    # "only" marker, not a minus sign.
    prefix = text[: match.start()]
    negative = "(" in prefix or "-" in prefix

    try:
        value = Decimal(digits)
    except InvalidOperation:
        return None
    return -value if negative else value


def _to_date(raw: object, *, dayfirst: bool) -> date | None:
    """Parse a date cell however the export wrote it: a real date or datetime,
    an Excel serial that lost its format, or text in any common notation."""
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, pd.Timestamp):
        return None if pd.isna(raw) else raw.date()
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float, Decimal)):
        if isinstance(raw, float) and pd.isna(raw):
            return None
        serial = float(raw)
        if _SERIAL_MIN <= serial <= _SERIAL_MAX:
            return (_EXCEL_EPOCH + timedelta(days=serial)).date()
        return None

    text = str(raw).strip()
    if not text:
        return None
    if text.isdigit() and len(text) == 5:
        serial = int(text)
        if _SERIAL_MIN <= serial <= _SERIAL_MAX:
            return (_EXCEL_EPOCH + timedelta(days=serial)).date()
        return None
    parsed = pd.to_datetime(
        text, errors="coerce", dayfirst=False if _ISO_DATE.match(text) else dayfirst
    )
    if parsed is pd.NaT or pd.isna(parsed):
        return None
    return parsed.date()


def _is_total_row(cells: list[object], mapping: dict[str, int]) -> bool:
    candidates = [cells[0] if cells else ""]
    for name in ("sales_row_id", "customer_name", "product"):
        if name in mapping:
            candidates.append(cells[mapping[name]])
    return any(_TOTAL_ROW.match(_text(candidate)) for candidate in candidates)


def _read_table(
    table: _Table,
    header_index: int,
    mapping: dict[str, int],
    *,
    document_id: str,
    filename: str,
    dayfirst: bool,
    currency: str,
    warnings: list[str],
) -> tuple[list[SalesRecord], SourceReadReport]:
    header = [_text(cell) for cell in table.rows[header_index]]
    # The header's width as it came off the file — the padded row is as wide
    # as the widest row, which is exactly the ragged row this must catch.
    width = (
        table.raw_widths[header_index]
        if header_index < len(table.raw_widths)
        else len(header)
    )
    columns = {
        canonical: header[position] or f"column {position + 1}"
        for canonical, position in mapping.items()
    }
    derived = "amount" not in mapping

    records: list[SalesRecord] = []
    skipped: dict[str, int] = {}
    filled: dict[str, int] = {}

    def skip(reason: str) -> None:
        skipped[reason] = skipped.get(reason, 0) + 1

    data_rows = table.rows[header_index + 1 :]
    for offset, cells in enumerate(data_rows):
        # Spreadsheet row as a human sees it: the header is row header_index+1.
        row_number = header_index + 2 + offset

        raw_width = (
            table.raw_widths[header_index + 1 + offset]
            if header_index + 1 + offset < len(table.raw_widths)
            else len(cells)
        )
        if raw_width > width and any(_text(cell) for cell in cells[width:]):
            raise SalesReadError(
                f"row {row_number} of the sales data has {raw_width} fields but "
                f"the header has {width}. If a value contains the delimiter — "
                'an amount like Rs. 45,900 — the field must be quoted: "Rs. 45,900"'
            )

        if not any(_text(cell) for cell in cells):
            skip("blank")
            continue
        if _is_total_row(cells, mapping):
            skip("total_row")
            continue

        when = _to_date(cells[mapping["date"]], dayfirst=dayfirst)
        if when is None:
            skip("no_date")
            continue

        if derived:
            quantity = _to_decimal(cells[mapping["quantity"]])
            price = _to_decimal(cells[mapping["unit_price"]])
            amount = quantity * price if quantity is not None and price is not None else None
        else:
            amount = _to_decimal(cells[mapping["amount"]])
        if amount is None:
            skip("no_amount")
            continue

        customer = _text(cells[mapping["customer_name"]]) if "customer_name" in mapping else ""
        if not customer:
            customer = UNSPECIFIED
            filled["customer_name"] = filled.get("customer_name", 0) + 1
        product = _text(cells[mapping["product"]]) if "product" in mapping else ""
        if not product:
            product = UNSPECIFIED
            filled["product"] = filled.get("product", 0) + 1
        region = _text(cells[mapping["region"]]) if "region" in mapping else ""
        row_id = _text(cells[mapping["sales_row_id"]]) if "sales_row_id" in mapping else ""

        records.append(
            SalesRecord(
                sales_row_id=row_id or f"SAL-{row_number:04d}",
                date=when,
                amount=amount,
                customer_name=customer,
                product=product,
                region=region or None,
                currency=currency,
                source=Provenance(document_id=document_id, row_number=row_number),
            )
        )

    report = SourceReadReport(
        document_id=document_id,
        filename=filename,
        format=table.format,
        sheet=table.sheet,
        encoding=table.encoding,
        delimiter=table.delimiter,
        header_row=header_index + 1,
        columns=columns,
        amount_derived=derived,
        rows_seen=len(data_rows),
        rows_used=len(records),
        rows_skipped=sum(skipped.values()),
        skipped=skipped,
        filled_defaults=filled,
        warnings=list(warnings),
    )
    return records, report


def read_sales_export(
    document_id: str,
    filename: str,
    content: bytes,
    *,
    dayfirst: bool = True,
    currency: str = "PKR",
) -> tuple[list[SalesRecord], SourceReadReport]:
    """Read a sales export of any supported format into `SalesRecord` objects,
    with a report of everything the reader decided along the way.

    Args:
        document_id: The stored upload this export came from.
        filename: Picks the reader by suffix — see `SUPPORTED_SUFFIXES`.
        content: The raw file bytes.
        dayfirst: Parse ambiguous dates as DD/MM/YYYY. True by default, which is
            the Pakistani convention — `03/06/2026` is 3 June, not 6 March.
        currency: ISO code recorded on every row.

    Returns:
        The usable rows in file order, and the `SourceReadReport` saying which
        sheet and header row they came from, how the columns were mapped, and
        which rows were skipped and why.

    Raises:
        SalesReadError: The format is unsupported, no sheet has a header naming
            a date and an amount (or quantity and price), the file is
            malformed, or no usable rows were found.
    """
    suffix = _suffix(filename)
    if suffix not in SUPPORTED_SUFFIXES:
        raise SalesReadError(
            f"unsupported sales data format: {filename!r}. Expected one of "
            f"{', '.join(sorted(SUPPORTED_SUFFIXES))}."
        )

    tables = _load_tables(content, filename, suffix)
    if not any(table.rows for table in tables):
        raise SalesReadError("the sales data has no rows")

    warnings: list[str] = []
    last_report: SourceReadReport | None = None
    for table in tables:
        found = _find_header(table.rows)
        if found is None:
            if table.sheet is not None:
                warnings.append(
                    f"sheet {table.sheet!r} has no header naming a date and an "
                    "amount; skipped"
                )
            continue
        header_index, mapping = found
        records, report = _read_table(
            table,
            header_index,
            mapping,
            document_id=document_id,
            filename=filename,
            dayfirst=dayfirst,
            currency=currency,
            warnings=warnings,
        )
        if records:
            if report.rows_skipped:
                logger.info(
                    "Sales data %s: used %s row(s), skipped %s (%s).",
                    document_id, report.rows_used, report.rows_skipped, report.skipped,
                )
            return records, report
        last_report = report
        if table.sheet is not None:
            warnings.append(f"sheet {table.sheet!r} has a header but no usable rows; skipped")

    if last_report is not None:
        raise SalesReadError(
            "no usable rows in the sales data: every row was missing a date or "
            f"an amount (skipped: {last_report.skipped})"
        )
    first_row = next(
        (
            [_text(cell) for cell in table.rows[0] if _text(cell)]
            for table in tables
            if table.rows
        ),
        [],
    )
    raise SalesReadError(
        "could not find a header row naming a date and an amount (or a quantity "
        f"and a unit price). The first row reads: {', '.join(first_row) or '(empty)'}"
    )


def read_sales_data(
    document_id: str,
    filename: str,
    content: bytes,
    *,
    dayfirst: bool = True,
    currency: str = "PKR",
) -> list[SalesRecord]:
    """`read_sales_export` for callers that only want the records."""
    records, _report = read_sales_export(
        document_id, filename, content, dayfirst=dayfirst, currency=currency
    )
    return records


# --------------------------------------------------------------------------- #
# 2. The analysis
# --------------------------------------------------------------------------- #


def _share(revenue: Decimal, total: Decimal) -> float:
    """Percent of the total, or 0.0 when the total makes the ratio meaningless."""
    if total <= 0:
        return 0.0
    return round(float(Decimal(100) * revenue / total), 2)


def _ranked(frame: pd.DataFrame, key: str) -> list[tuple[str, Decimal, int]]:
    """Group revenue by `key`, highest revenue first, then by name.

    The sort is what makes two runs over the same records produce the same
    list; the schema rejects a breakdown that arrives in any other order.
    """
    rows = [
        (str(name), sum(group["amount"], Decimal(0)), len(group))
        for name, group in frame.groupby(key, sort=False)
    ]
    rows.sort(key=lambda row: (-row[1], row[0]))
    return rows


def _negative_amounts(records: list[SalesRecord]) -> list[dict]:
    """A sale with a negative amount: a refund, a correction, or a typo."""
    return [
        {
            "kind": "negative-amount",
            "row_id": record.sales_row_id,
            "month": None,
            "related": [],
            "explanation": (
                f"sale of {_money(record.amount)} to {record.customer_name} on "
                f"{record.date.isoformat()} is negative — a refund, a correction, "
                "or a sign error in the export"
            ),
        }
        for record in records
        if record.amount < 0
    ]


def _duplicate_transactions(frame: pd.DataFrame) -> list[dict]:
    """The same date, customer, product, and amount, more than once.

    Genuine repeat purchases do land on the same key, which is why this is a
    finding for a human and not a verdict: the explanation says how many rows
    share the key and names them all.
    """
    findings: list[dict] = []
    if frame.empty:
        return findings
    keyed = frame.groupby(
        ["date", "customer_name", "product", "amount"], sort=False, dropna=False
    )
    for _key, group in keyed:
        if len(group) < 2:
            continue
        ids = list(group["sales_row_id"])
        first = group.iloc[0]
        findings.append(
            {
                "kind": "duplicate-transaction",
                "row_id": ids[0],
                "month": None,
                "related": ids,
                "explanation": (
                    f"{len(ids)} identical rows — {first['customer_name']}, "
                    f"{first['product']}, {_money(first['amount'])} on "
                    f"{first['date'].isoformat()} — share one key"
                ),
            }
        )
    return findings


def _revenue_spikes(monthly: list[MonthlyRevenue]) -> list[dict]:
    """A month far from the median month: more than double it, or under half.

    Against the median rather than the mean so one enormous month cannot lift
    the bar high enough to hide behind. Months are compared only when there
    are at least `SPIKE_MIN_MONTHS` of them.
    """
    if len(monthly) < SPIKE_MIN_MONTHS:
        return []
    median = statistics.median([entry.revenue for entry in monthly])
    if median <= 0:
        return []

    findings: list[dict] = []
    for entry in monthly:
        if entry.revenue > median * SPIKE_FACTOR:
            findings.append(
                {
                    "kind": "revenue-spike",
                    "row_id": None,
                    "month": entry.month,
                    "related": [],
                    "explanation": (
                        f"{entry.month} revenue of {_money(entry.revenue)} is more "
                        f"than double the median month ({_money(median)})"
                    ),
                }
            )
        elif entry.revenue < median / SPIKE_FACTOR:
            findings.append(
                {
                    "kind": "revenue-spike",
                    "row_id": None,
                    "month": entry.month,
                    "related": [],
                    "explanation": (
                        f"{entry.month} revenue of {_money(entry.revenue)} is less "
                        f"than half the median month ({_money(median)})"
                    ),
                }
            )
    return findings


def _large_transactions(records: list[SalesRecord]) -> list[dict]:
    """A sale far above the median sale: worth a look, not a conclusion."""
    if len(records) < LARGE_TRANSACTION_MIN_SAMPLE:
        return []
    median = statistics.median([record.amount for record in records])
    if median <= 0:
        return []
    return [
        {
            "kind": "large-transaction",
            "row_id": record.sales_row_id,
            "month": None,
            "related": [],
            "explanation": (
                f"{record.sales_row_id}: {_money(record.amount)} from "
                f"{record.customer_name} is more than {LARGE_TRANSACTION_FACTOR}× "
                f"the median sale ({_money(median)})"
            ),
        }
        for record in records
        if record.amount > median * LARGE_TRANSACTION_FACTOR
    ]


def _money(value: Decimal) -> str:
    return f"{value:,.2f}"


def analyze_sales(
    records: list[SalesRecord],
    reports: Sequence[SourceReadReport] = (),
) -> SalesAnalyticsResult:
    """Aggregate sales records into the full analytics readout.

    Pure arithmetic: pandas groups the rows, Decimal keeps the money exact, and
    the breakdowns partition the records — by month and by product they each
    sum back to `total_revenue` and `record_count`, which the schema enforces.

    Args:
        records: The sales rows, typically from every export in a case
            concatenated in upload order.
        reports: The `SourceReadReport` of each export the records came from,
            carried into the readout so it can say how the data was cleaned.

    Returns:
        A `SalesAnalyticsResult`. Empty input yields an empty readout rather
        than an error — the caller decides whether no sales data is exceptional.
    """
    generated_at = datetime.now(timezone.utc)
    if not records:
        return SalesAnalyticsResult(
            record_count=0,
            total_revenue=Decimal(0),
            document_ids=[],
            data_quality=list(reports),
            generated_at=generated_at,
        )

    frame = pd.DataFrame(
        {
            "sales_row_id": [record.sales_row_id for record in records],
            "date": [record.date for record in records],
            "customer_name": [record.customer_name for record in records],
            "product": [record.product for record in records],
            "region": [record.region for record in records],
            "amount": [record.amount for record in records],
        }
    )
    # Lexicographic order on YYYY-MM is chronological order, which is the order
    # the schema requires of monthly_revenue.
    frame["month"] = frame["date"].map(lambda day: f"{day.year:04d}-{day.month:02d}")

    total = sum((record.amount for record in records), Decimal(0))

    monthly = [
        MonthlyRevenue(
            month=month,
            revenue=sum(group["amount"], Decimal(0)),
            transaction_count=len(group),
        )
        for month, group in frame.groupby("month", sort=True)
    ]
    products = [
        ProductRevenue(
            product=name,
            revenue=revenue,
            transaction_count=count,
            share=_share(revenue, total),
        )
        for name, revenue, count in _ranked(frame, "product")
    ]
    # dropna is the default: rows without a region are simply not in this
    # breakdown, rather than being filed under a label they never carried.
    regions = [
        RegionSummary(
            region=name,
            revenue=revenue,
            transaction_count=count,
            share=_share(revenue, total),
        )
        for name, revenue, count in _ranked(frame, "region")
    ]
    customers = [
        CustomerSummary(
            customer_name=name,
            revenue=revenue,
            transaction_count=count,
            share=_share(revenue, total),
        )
        for name, revenue, count in _ranked(frame, "customer_name")
    ][:TOP_CUSTOMERS]

    records_by_id = {record.sales_row_id: record for record in records}
    findings = (
        _negative_amounts(records)
        + _duplicate_transactions(frame)
        + _revenue_spikes(monthly)
        + _large_transactions(records)
    )
    anomalies = [
        Anomaly(
            anomaly_id=f"ANM-{position:04d}",
            kind=finding["kind"],
            explanation=finding["explanation"],
            source_row_id=finding["row_id"],
            related_row_ids=list(finding["related"]),
            month=finding["month"],
            source=(
                records_by_id[finding["row_id"]].source
                if finding["row_id"] in records_by_id
                else None
            ),
        )
        for position, finding in enumerate(findings, start=1)
    ]

    return SalesAnalyticsResult(
        record_count=len(records),
        period_start=min(record.date for record in records),
        period_end=max(record.date for record in records),
        total_revenue=total,
        monthly_revenue=monthly,
        revenue_by_product=products,
        top_customers=customers,
        sales_by_region=regions,
        anomalies=anomalies,
        document_ids=sorted({record.source.document_id for record in records}),
        data_quality=list(reports),
        generated_at=generated_at,
    )


# --------------------------------------------------------------------------- #
# 3. Exporting the readout
# --------------------------------------------------------------------------- #

_MONEY_FORMAT = "#,##0.00"


def _sheet(workbook: Workbook, title: str, header: list[str], rows: list[list[object]]) -> None:
    """One worksheet: a bold header, the rows, money formatted, columns sized."""
    sheet = workbook.create_sheet(title=title)
    sheet.append(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(row)
    for column_index, name in enumerate(header, start=1):
        letter = get_column_letter(column_index)
        longest = max(
            [len(str(name))] + [len(str(row[column_index - 1])) for row in rows if row[column_index - 1] is not None],
            default=len(str(name)),
        )
        sheet.column_dimensions[letter].width = min(max(12, longest + 2), 60)
        if name.lower().startswith(("revenue", "amount", "total revenue")):
            for cell in sheet[letter][1:]:
                cell.number_format = _MONEY_FORMAT
    sheet.freeze_panes = "A2"


def export_workbook(result: SalesAnalyticsResult) -> bytes:
    """Render a saved readout as an Excel workbook: one sheet per breakdown,
    the anomalies, and the data-quality report of every export it was read
    from. Every figure is copied from the readout; nothing is recomputed."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    period = (
        f"{result.period_start.isoformat()} to {result.period_end.isoformat()}"
        if result.period_start and result.period_end
        else "—"
    )
    _sheet(
        workbook,
        "Summary",
        ["Measure", "Value"],
        [
            ["Sales records", result.record_count],
            ["Period", period],
            ["Total revenue", result.total_revenue],
            ["Months", len(result.monthly_revenue)],
            ["Products", len(result.revenue_by_product)],
            ["Regions", len(result.sales_by_region)],
            ["Anomalies", len(result.anomalies)],
            ["Read from", ", ".join(result.document_ids) or "—"],
            ["Generated at (UTC)", result.generated_at.isoformat()],
            ["Note", "Anomalies are findings for a human, never verdicts."],
        ],
    )
    _sheet(
        workbook,
        "Monthly revenue",
        ["Month", "Revenue", "Transactions"],
        [[entry.month, entry.revenue, entry.transaction_count] for entry in result.monthly_revenue],
    )
    _sheet(
        workbook,
        "By product",
        ["Product", "Revenue", "Transactions", "Share %"],
        [
            [entry.product, entry.revenue, entry.transaction_count, entry.share]
            for entry in result.revenue_by_product
        ],
    )
    _sheet(
        workbook,
        "By region",
        ["Region", "Revenue", "Transactions", "Share %"],
        [
            [entry.region, entry.revenue, entry.transaction_count, entry.share]
            for entry in result.sales_by_region
        ],
    )
    _sheet(
        workbook,
        "Top customers",
        ["Customer", "Revenue", "Transactions", "Share %"],
        [
            [entry.customer_name, entry.revenue, entry.transaction_count, entry.share]
            for entry in result.top_customers
        ],
    )
    _sheet(
        workbook,
        "Anomalies",
        ["Id", "Kind", "Row", "Month", "Related rows", "Explanation"],
        [
            [
                anomaly.anomaly_id,
                anomaly.kind,
                anomaly.source_row_id or "",
                anomaly.month or "",
                ", ".join(anomaly.related_row_ids),
                anomaly.explanation,
            ]
            for anomaly in result.anomalies
        ],
    )
    _sheet(
        workbook,
        "Data quality",
        [
            "File", "Format", "Sheet", "Encoding", "Header row", "Rows seen",
            "Rows used", "Rows skipped", "Skipped (why)", "Filled as Unspecified",
            "Amount derived", "Columns mapped", "Warnings",
        ],
        [
            [
                report.filename,
                report.format,
                report.sheet or "",
                report.encoding or "",
                report.header_row,
                report.rows_seen,
                report.rows_used,
                report.rows_skipped,
                ", ".join(f"{reason}: {count}" for reason, count in report.skipped.items()),
                ", ".join(f"{name}: {count}" for name, count in report.filled_defaults.items()),
                "yes (quantity × unit price)" if report.amount_derived else "no",
                ", ".join(f"{canonical} ← {source}" for canonical, source in report.columns.items()),
                "; ".join(report.warnings),
            ]
            for report in result.data_quality
        ],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
