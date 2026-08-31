"""Reports: the content, the two files, the immutable history, and the route.

The content builder is tested as data — decided items in, pending items
named but not listed — and the renderers are checked for producing a real PDF
and a real workbook that carry that content. The route tests prove a report
lands in storage, in the `reports` table, and in the audit trail, and that the
table refuses to be rewritten.
"""

from __future__ import annotations

import io
import zipfile
from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.sqlite_store import ReportImmutable, SqliteCaseRepository
from app.modules.reports import service as reports
from app.shared.schemas import AuditAction, ReportRecord, ReviewDecision
from tests.conftest import DEMO_ORG_ID, DEMO_USER, OTHER_ORG_ID, load_sample_queue

GENERATED_AT = datetime(2026, 8, 29, 10, 0, tzinfo=timezone.utc)


def a_report(repository: SqliteCaseRepository, case_id: str):
    case = repository.get_case(DEMO_ORG_ID, case_id)
    items = repository.list_review_items(DEMO_ORG_ID, case_id)
    audit = repository.list_audit(DEMO_ORG_ID, case_id)
    benford = repository.get_benford(DEMO_ORG_ID, case_id)
    return reports.generate_report(
        case, items, audit, benford,
        report_id="RPT-test000001", generated_by=DEMO_USER.user_id, generated_at=GENERATED_AT,
    )


# --------------------------------------------------------------------------- #
# Content
# --------------------------------------------------------------------------- #


def test_only_decided_items_are_reported_as_findings(repository, seeded_case: str) -> None:
    files = a_report(repository, seeded_case)
    content = files.content
    decided = content.sections[0]
    assert decided.title == "Decided items"

    queue = load_sample_queue().items
    expected = {item.review_item_id for item in queue if item.decision is not ReviewDecision.PENDING}
    listed = {row[0] for row in decided.rows}
    assert listed == expected
    assert content.pending_count == sum(1 for item in queue if item.decision is ReviewDecision.PENDING)
    # Pending items are named as pending, never as findings.
    assert decided.note is not None and "awaiting a decision are excluded" in decided.note
    for item_id in content.excluded_pending_items:
        assert item_id in decided.note


def test_the_summary_states_the_counts_and_the_period(repository, seeded_case: str) -> None:
    summary = dict(a_report(repository, seeded_case).content.summary)
    assert summary["Client"] == "Haroon Textiles"
    assert summary["Case"] == seeded_case
    assert summary["Period covered"] == "2026-06-02 to 2026-06-18"
    assert summary["Ledger rows reviewed"] == "10"
    assert summary["Reconciliation"] == "8 matched, 1 partial, 1 unmatched"
    assert summary["Human decisions"].startswith("1 approved, 1 rejected, 8 still pending")
    assert summary["Report id"] == "RPT-test000001"


def test_flags_provenance_benford_and_the_trail_are_all_present(repository, seeded_case: str) -> None:
    titles = [section.title for section in a_report(repository, seeded_case).content.sections]
    assert titles == [
        "Decided items",
        "Red flags on decided items",
        "Provenance",
        "Benford's law: first-digit distribution",
        "Audit trail",
    ]


def test_provenance_names_the_document_and_the_location(repository, seeded_case: str) -> None:
    provenance = a_report(repository, seeded_case).content.sections[2]
    # RI-0001 is approved: its ledger row, its bank line, and its AI readings.
    rows = [row for row in provenance.rows if row[0] == "RI-0001"]
    assert ["Ledger row", "DOC-LED-001", "row 5"] == rows[0][1:4]
    assert ["Bank statement", "DOC-BNK-001", "page 1"] == rows[1][1:4]
    assert any(row[1].startswith("AI reading") and "AI confidence" in row[5] for row in rows)


def test_nothing_is_recomputed(repository, seeded_case: str) -> None:
    """The amounts in the report are the ledger's, character for character."""
    content = a_report(repository, seeded_case).content
    amounts = {row[3] for row in content.sections[0].rows}
    assert amounts == {"PKR 284,000.00", "PKR 45,900.00"}


def test_the_record_carries_digests_of_exactly_the_files(repository, seeded_case: str) -> None:
    import hashlib

    files = a_report(repository, seeded_case)
    assert files.record.pdf_sha256 == hashlib.sha256(files.pdf).hexdigest()
    assert files.record.excel_sha256 == hashlib.sha256(files.excel).hexdigest()
    assert files.record.item_count == 10
    assert files.record.approved_count + files.record.rejected_count + files.record.pending_count == 10


# --------------------------------------------------------------------------- #
# The files
# --------------------------------------------------------------------------- #


def test_the_pdf_is_a_pdf_with_the_expected_pages(repository, seeded_case: str) -> None:
    import pymupdf

    files = a_report(repository, seeded_case)
    assert files.pdf.startswith(b"%PDF-")
    with pymupdf.open(stream=files.pdf, filetype="pdf") as document:
        text = "".join(page.get_text() for page in document)
        assert document.page_count >= 6  # summary + five sections
    assert "Tarazu" in text and "Haroon Textiles" in text
    assert "Decided items" in text and "Audit trail" in text
    assert "RI-0001" in text and "RI-0004" in text
    assert "The AI suggests, the human decides." in text


def test_the_workbook_carries_one_sheet_per_section(repository, seeded_case: str) -> None:
    files = a_report(repository, seeded_case)
    assert zipfile.is_zipfile(io.BytesIO(files.excel))
    workbook = load_workbook(io.BytesIO(files.excel), read_only=True)
    assert workbook.sheetnames[0] == "Summary"
    assert "Decided items" in workbook.sheetnames
    assert "Audit trail" in workbook.sheetnames
    decided = workbook["Decided items"]
    flat = {
        str(value)
        for row in decided.iter_rows(values_only=True)
        for value in row
        if value is not None
    }
    assert "RI-0001" in flat and "PKR 284,000.00" in flat


def test_rendering_is_deterministic(repository, seeded_case: str) -> None:
    """Same inputs, same workbook — and the same report content for the PDF."""
    first = a_report(repository, seeded_case)
    second = a_report(repository, seeded_case)
    assert first.content == second.content
    assert first.excel == second.excel


# --------------------------------------------------------------------------- #
# The history is append-only
# --------------------------------------------------------------------------- #


def a_record(case_id: str, report_id: str = "RPT-a") -> ReportRecord:
    return ReportRecord(
        report_id=report_id, case_id=case_id, generated_by=DEMO_USER.user_id,
        generated_at=GENERATED_AT, pdf_path="x.pdf", excel_path="x.xlsx",
        pdf_sha256="a" * 64, excel_sha256="b" * 64,
        item_count=3, approved_count=1, rejected_count=1, pending_count=1,
        flag_count=2, audit_record_count=9,
    )


def test_the_reports_table_refuses_update_and_delete(repository, seeded_case: str) -> None:
    repository.save_report(DEMO_ORG_ID, a_record(seeded_case))
    with pytest.raises(ReportImmutable):
        repository._write([("update reports set pdf_path = 'tampered' where report_id = 'RPT-a'", ())])
    with pytest.raises(ReportImmutable):
        repository._write([("delete from reports where report_id = 'RPT-a'", ())])
    assert repository.get_report(DEMO_ORG_ID, "RPT-a").pdf_path == "x.pdf"


def test_a_report_id_is_never_reused(repository, seeded_case: str) -> None:
    repository.save_report(DEMO_ORG_ID, a_record(seeded_case))
    with pytest.raises(Exception):
        repository.save_report(DEMO_ORG_ID, a_record(seeded_case))


def test_reports_are_listed_newest_first_and_scoped(repository, seeded_case: str, other_org: str) -> None:
    repository.save_report(DEMO_ORG_ID, a_record(seeded_case, "RPT-first"))
    later = a_record(seeded_case, "RPT-second").model_copy(
        update={"generated_at": GENERATED_AT.replace(hour=11)}
    )
    repository.save_report(DEMO_ORG_ID, later)
    assert [r.report_id for r in repository.list_reports(DEMO_ORG_ID, seeded_case)] == [
        "RPT-second", "RPT-first"
    ]
    assert repository.list_reports(OTHER_ORG_ID, seeded_case) == []
    assert repository.get_report(OTHER_ORG_ID, "RPT-first") is None


def test_the_schema_refuses_counts_that_do_not_add_up() -> None:
    with pytest.raises(ValueError, match="must equal item_count"):
        ReportRecord.model_validate({**a_record("CASE-X").model_dump(), "pending_count": 5})


# --------------------------------------------------------------------------- #
# The route
# --------------------------------------------------------------------------- #


def test_generating_a_report_stores_it_records_it_and_logs_it(
    client: TestClient, repository, storage, seeded_case: str
) -> None:
    before = len(repository.list_audit(DEMO_ORG_ID, seeded_case))
    response = client.post("/v1/reports", json={})
    assert response.status_code == 201
    body = response.json()
    assert body["case_id"] == seeded_case
    assert body["item_count"] == 10
    assert body["pending_count"] == 8
    assert body["downloads"]["pdf"].endswith("?format=pdf")

    record = repository.get_report(DEMO_ORG_ID, body["report_id"])
    assert record is not None
    assert storage.get(record.pdf_path).startswith(b"%PDF-")
    assert storage.get(record.excel_path).startswith(b"PK")

    trail = repository.list_audit(DEMO_ORG_ID, seeded_case)
    assert len(trail) == before + 1
    assert trail[-1].action is AuditAction.REPORT_GENERATED
    assert trail[-1].item_id == body["report_id"]
    assert trail[-1].actor_id == DEMO_USER.user_id


def test_the_files_download_with_the_right_types(client: TestClient, seeded_case: str) -> None:
    report_id = client.post("/v1/reports", json={}).json()["report_id"]

    pdf = client.get(f"/v1/reports/{report_id}/download", params={"format": "pdf"})
    assert pdf.status_code == 200
    assert pdf.headers["content-type"] == "application/pdf"
    assert pdf.content.startswith(b"%PDF-")
    assert f'filename="tarazu-{seeded_case}-{report_id}.pdf"' in pdf.headers["content-disposition"]

    excel = client.get(f"/v1/reports/{report_id}/download", params={"format": "excel"})
    assert excel.status_code == 200
    assert excel.content.startswith(b"PK")
    assert excel.headers["content-disposition"].endswith('.xlsx"')


def test_the_history_lists_every_generation(client: TestClient, seeded_case: str) -> None:
    first = client.post("/v1/reports", json={}).json()["report_id"]
    second = client.post("/v1/reports", json={"case_id": seeded_case}).json()["report_id"]
    listed = client.get("/v1/reports").json()
    assert listed["total"] == 2
    assert [r["report_id"] for r in listed["reports"]] == [second, first]


def test_another_firm_cannot_see_or_download_the_report(
    client: TestClient, other_client: TestClient, seeded_case: str
) -> None:
    report_id = client.post("/v1/reports", json={}).json()["report_id"]
    assert other_client.get(f"/v1/reports/{report_id}/download").status_code == 404
    assert other_client.post("/v1/reports", json={"case_id": seeded_case}).status_code == 404


def test_a_report_needs_a_case(client: TestClient, demo_org: str) -> None:
    assert client.post("/v1/reports", json={}).status_code == 404


def test_a_report_after_new_decisions_is_a_new_record(
    client: TestClient, repository, seeded_case: str
) -> None:
    first = client.post("/v1/reports", json={}).json()
    pending = next(
        item for item in repository.list_review_items(DEMO_ORG_ID, seeded_case)
        if item.decision is ReviewDecision.PENDING
    )
    client.post(f"/v1/review-items/{pending.review_item_id}/approve", json={})
    second = client.post("/v1/reports", json={}).json()

    assert second["report_id"] != first["report_id"]
    assert second["approved_count"] == first["approved_count"] + 1
    # The first report is untouched and still downloadable.
    assert client.get(f"/v1/reports/{first['report_id']}/download").status_code == 200
