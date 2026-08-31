"""What Phase 1 added to the deliverable: letterhead, corrections, Urdu, sign-off.

Each of these is additive. A one-off engagement for a firm that has filled in
no branding produces exactly the report it always did — the tests for that live
in `test_reports.py` and still pass unchanged. What is pinned here is that when
the extra facts exist, they reach the file, and that none of them touches a
figure.
"""

from __future__ import annotations

from datetime import datetime, timezone

import pytest

from app.core.repository import StoredDocument
from app.core.sqlite_store import SqliteCaseRepository
from app.modules.reports import service as reports
from app.modules.reports.content import ReportBranding
from app.shared.schemas import (
    AssistantLanguage,
    CaseRecord,
    Client,
    ClientRuleConfig,
    DocumentType,
    OrgProfile,
    ReviewDecision,
)
from tests.conftest import DEMO_ORG_ID, load_sample_queue

A_PNG = (
    # A 1x1 transparent PNG: enough for the renderer to size and place.
    "data:image/png;base64,"
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk"
    "YPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
)


def _case() -> CaseRecord:
    return CaseRecord(
        case_id="CASE-REPORT",
        client_name="Haroon Textiles",
        created_by="auditor",
        created_at=datetime(2026, 6, 30, 12, 0, tzinfo=timezone.utc),
    )


def _generate(**kwargs):
    queue = load_sample_queue()
    return reports.generate_report(
        _case(),
        queue.items,
        [],
        None,
        report_id="RPT-test",
        generated_by="auditor",
        generated_at=datetime(2026, 6, 30, 12, 0, tzinfo=timezone.utc),
        **kwargs,
    )


# --------------------------------------------------------------------------- #
# Letterhead
# --------------------------------------------------------------------------- #


def test_branding_reaches_the_content_and_both_files() -> None:
    branding = ReportBranding(
        firm_name="Tarazu Demo Firm",
        legal_name="Lahore Audit Associates",
        address="12 Gulberg III, Lahore",
        registration_number="ICAP-1234",
        logo=A_PNG,
        footer="Prepared under ISA 500.",
    )
    files = _generate(branding=branding)

    assert files.content.branding is branding
    assert files.content.branding.display_name == "Lahore Audit Associates"
    assert "ICAP-1234" in files.content.branding.contact_line
    assert files.pdf[:4] == b"%PDF"
    assert files.excel[:2] == b"PK"


def test_a_malformed_logo_does_not_lose_the_report() -> None:
    """The deliverable is the reconciliation, not the picture."""
    files = _generate(
        branding=ReportBranding(firm_name="A Firm", logo="data:image/png;base64,not-base64!!")
    )
    assert files.pdf[:4] == b"%PDF"


def test_no_branding_still_renders() -> None:
    assert _generate().pdf[:4] == b"%PDF"


# --------------------------------------------------------------------------- #
# Corrections
# --------------------------------------------------------------------------- #


def test_corrections_appear_as_their_own_section_with_both_readings(
    repository: SqliteCaseRepository,
) -> None:
    from app.shared.schemas import ValueCorrection

    correction = ValueCorrection(
        correction_id="COR-1",
        case_id="CASE-REPORT",
        review_item_id="RI-0001",
        document_id="DOC-BNK-1",
        field="amount",
        ai_value="49500.00",
        corrected_value="49900.00",
        note="Smudged digit.",
        corrected_by="auditor",
        corrected_at=datetime(2026, 6, 30, 11, 0, tzinfo=timezone.utc),
    )
    files = _generate(corrections=[correction])

    section = next(
        s for s in files.content.sections if s.title == "Corrections to extracted values"
    )
    row = section.rows[0]
    assert "49500.00" in row and "49900.00" in row, "both readings are kept"
    assert "did not re-run matching" in section.note
    assert any(label == "Corrections recorded" for label, _ in files.content.summary)


def test_a_report_without_corrections_has_no_such_section() -> None:
    titles = [section.title for section in _generate().content.sections]
    assert "Corrections to extracted values" not in titles


# --------------------------------------------------------------------------- #
# The Urdu executive summary
# --------------------------------------------------------------------------- #


def test_the_urdu_summary_is_composed_from_the_counts() -> None:
    files = _generate(urdu=True)
    summary = files.content.urdu_summary

    assert summary is not None
    assert "Haroon Textiles" in summary
    # The item count appears as a Western numeral, as the module documents.
    assert str(len(load_sample_queue().items)) in summary
    # And it is Urdu, not a translated placeholder.
    assert "جائزہ" in summary


def test_the_urdu_summary_is_off_by_default() -> None:
    assert _generate().content.urdu_summary is None


def test_the_workbook_carries_an_urdu_sheet() -> None:
    import io

    from openpyxl import load_workbook

    files = _generate(urdu=True)
    workbook = load_workbook(io.BytesIO(files.excel))
    assert "Urdu summary" in workbook.sheetnames
    sheet = workbook["Urdu summary"]
    assert sheet.sheet_view.rightToLeft is True
    assert "جائزہ" in sheet.cell(row=3, column=1).value


def test_reports_stay_byte_reproducible_with_the_extras() -> None:
    """The digest on the record must describe the report, not the second it ran."""
    branding = ReportBranding(firm_name="A Firm", logo=A_PNG, footer="Footer.")
    first = _generate(branding=branding, urdu=True)
    second = _generate(branding=branding, urdu=True)

    assert first.excel == second.excel
    assert first.record.excel_sha256 == second.record.excel_sha256


# --------------------------------------------------------------------------- #
# The sign-off gate, through the route
# --------------------------------------------------------------------------- #


@pytest.fixture()
def case_needing_sign_off(
    repository: SqliteCaseRepository, seeded_case: str, demo_org: str
) -> str:
    """The seeded case, attached to a client whose rules demand a signature."""
    client = Client(
        client_id="CLI-signoff",
        name="Haroon Textiles",
        rules=ClientRuleConfig(require_sign_off=True),
        language=AssistantLanguage.URDU,
        created_by="auditor",
        created_at=datetime.now(timezone.utc),
    )
    repository.create_client(demo_org, client)
    repository.update_case(
        demo_org,
        seeded_case,
        client_name=client.name,
        period_start=None,
        period_end=None,
        client_id=client.client_id,
    )
    return seeded_case


def test_a_report_is_refused_until_the_engagement_is_signed_off(
    client, case_needing_sign_off: str
) -> None:
    response = client.post("/v1/reports", json={"case_id": case_needing_sign_off})
    assert response.status_code == 409
    assert "requires a sign-off" in response.json()["detail"]


def test_a_signed_engagement_reports_in_the_clients_language(
    client, repository: SqliteCaseRepository, case_needing_sign_off: str, storage
) -> None:
    from app.core.auth import AuthenticatedUser
    from app.shared.schemas import OrgRole
    from tests.conftest import join, signed_in

    # Decide everything as the auditor, then sign off as a colleague.
    for item in client.get(f"/v1/review-items?case_id={case_needing_sign_off}").json()[
        "items"
    ]:
        if item["decision"] == ReviewDecision.PENDING.value:
            client.post(f"/v1/review-items/{item['review_item_id']}/approve", json={})

    partner = AuthenticatedUser(
        user_id="00000000-0000-4000-8000-0000000000f9", email="partner@tarazu.local"
    )
    join(repository, DEMO_ORG_ID, "Tarazu Demo Firm", partner, role=OrgRole.MEMBER)
    with signed_in(repository, storage, partner) as partner_client:
        signed = partner_client.post(
            "/v1/sign-offs", json={"case_id": case_needing_sign_off}
        )
        assert signed.status_code == 201, signed.text

    response = client.post("/v1/reports", json={"case_id": case_needing_sign_off})
    assert response.status_code == 201, response.text

    # The client reads Urdu, so the workbook carries the owner's summary.
    import io

    from openpyxl import load_workbook

    excel = client.get(
        f"/v1/reports/{response.json()['report_id']}/download?format=excel"
    )
    assert excel.status_code == 200
    assert "Urdu summary" in load_workbook(io.BytesIO(excel.content)).sheetnames


def test_a_case_with_no_client_needs_no_sign_off(client, seeded_case: str) -> None:
    """The gate is opt-in per client; the default firm is unaffected."""
    assert client.post("/v1/reports", json={"case_id": seeded_case}).status_code == 201


def test_generating_a_report_marks_the_case_reported(
    client, repository: SqliteCaseRepository, seeded_case: str
) -> None:
    from app.shared.schemas import CaseStatus

    client.post("/v1/reports", json={"case_id": seeded_case})
    assert repository.get_case(DEMO_ORG_ID, seeded_case).status is CaseStatus.REPORTED


def test_the_firms_letterhead_reaches_a_generated_report(
    client, repository: SqliteCaseRepository, seeded_case: str, demo_org: str
) -> None:
    repository.save_org_profile(
        OrgProfile(
            org_id=demo_org,
            legal_name="Lahore Audit Associates",
            registration_number="ICAP-1234",
            updated_at=datetime.now(timezone.utc),
        )
    )
    created = client.post("/v1/reports", json={"case_id": seeded_case})
    assert created.status_code == 201

    import io

    from openpyxl import load_workbook

    excel = client.get(
        f"/v1/reports/{created.json()['report_id']}/download?format=excel"
    )
    values = [
        cell.value
        for row in load_workbook(io.BytesIO(excel.content))["Summary"].iter_rows()
        for cell in row
    ]
    assert "Lahore Audit Associates" in values
