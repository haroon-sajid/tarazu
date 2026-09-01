"""Sales analytics tests. No network, no model — the module is pure pandas.

The reader tests feed real CSV / TSV / Excel / ODS / JSON bytes in the shapes
client exports actually arrive in; the analysis tests drive `analyze_sales`
over handcrafted records so every figure is checkable by hand; the route tests
run the real endpoints over the real in-memory store, end to end: upload an
export, run, read back, download.

Sales analytics is a standalone data source. The case pipeline (bank
statement, ledger, invoices) never runs it — the last test proves that.
"""

from __future__ import annotations

import ast
import io
import json
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.audit import Actor, ActorType
from app.core.config import DEFAULT_ORG_ID as ORG
from app.core.repository import StoredDocument
from app.core.sqlite_store import LocalDocumentStore, SqliteCaseRepository
from app.modules.analytics import service as analytics
from app.pipeline import run_pipeline
from app.shared.schemas import (
    AuditAction,
    CaseRecord,
    CaseStatus,
    DocumentType,
    Provenance,
    SalesDataUpload,
    SalesRecord,
)

USER_ID = "00000000-0000-4000-8000-000000000001"
AUDITOR = Actor.human(USER_ID)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #


def a_sale(
    row_id: str,
    day: date,
    customer: str,
    product: str,
    amount: str,
    region: str | None = None,
    document_id: str = "SLS-0001",
    row_number: int = 2,
) -> SalesRecord:
    return SalesRecord(
        sales_row_id=row_id,
        date=day,
        amount=Decimal(amount),
        customer_name=customer,
        product=product,
        region=region,
        source=Provenance(document_id=document_id, row_number=row_number),
    )


def a_sales_csv() -> bytes:
    """Three clean rows: June ×2 (with regions), July ×1 (without one).

    Money containing the delimiter is quoted, the way a real CSV export writes
    it (RFC 4180) — see the refusal test below for what happens when it is not.
    """
    return (
        "Sale Date,Customer,Item,Region,Amount\n"
        '02/06/2026,Gulberg Traders,Yarn,Punjab,"Rs. 45,900/-"\n'
        "10/06/2026,Al-Habib Stationers,Cloth,Sindh,12000\n"
        "15/07/2026,Gulberg Traders,Cloth,,30000\n"
    ).encode("utf-8")


def a_workbook(sheets: dict[str, list[list[object]]]) -> bytes:
    """An .xlsx with the given sheets, rows written exactly as passed."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(title=name)
        for row in rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def a_ledger() -> bytes:
    buffer = io.BytesIO()
    pd.DataFrame(
        {
            "Date": ["02/06/2026", "10/06/2026", "14/06/2026"],
            "Party Name": [
                "Gulberg Traders (Pvt) Ltd",
                "Al-Habib Stationers",
                "Indus Power Solutions",
            ],
            "Amount": [284000, 45900, 1500000],
            "Particulars": ["Yarn purchase", "Office supplies", "Generator advance"],
        }
    ).to_excel(buffer, index=False)
    return buffer.getvalue()


def create_case(repository: SqliteCaseRepository, org_id: str, case_id: str) -> str:
    repository.create_case(
        org_id,
        CaseRecord(
            case_id=case_id,
            client_name="Haroon Textiles",
            status=CaseStatus.READY_FOR_REVIEW,
            created_by=USER_ID,
            created_at=datetime.now(timezone.utc),
        ),
    )
    return case_id


def seed_sales_upload(
    repository: SqliteCaseRepository,
    org_id: str,
    storage: LocalDocumentStore,
    case_id: str,
    sales_data_id: str = "SLS-0001",
    filename: str = "sales.csv",
    content: bytes | None = None,
) -> SalesDataUpload:
    """Put a sales export in storage and on the case's uploads table, the way
    the upload route does — without going through HTTP."""
    content = a_sales_csv() if content is None else content
    upload = SalesDataUpload(
        sales_data_id=sales_data_id,
        org_id=org_id,
        case_id=case_id,
        filename=filename,
        size_bytes=len(content),
        storage_path=f"{case_id}/sales-data/{sales_data_id}/{filename}",
        uploaded_by=USER_ID,
        uploaded_at=datetime.now(timezone.utc),
    )
    storage.put(upload.storage_path, content, "text/csv")
    repository.add_sales_data_upload(org_id, case_id, upload)
    return upload


def seed_sales_case(
    repository: SqliteCaseRepository,
    org_id: str,
    storage: LocalDocumentStore,
    case_id: str = "CASE-SLS-001",
) -> str:
    """A case in `org_id` carrying one uploaded sales export."""
    create_case(repository, org_id, case_id)
    seed_sales_upload(repository, org_id, storage, case_id)
    return case_id


def upload_sales(client: TestClient, case_id: str, filename: str, content: bytes):
    return client.post(
        f"/v1/cases/{case_id}/sales-data",
        files={"file": (filename, io.BytesIO(content), "application/octet-stream")},
    )


# --------------------------------------------------------------------------- #
# Reading the export
# --------------------------------------------------------------------------- #


def test_reading_a_csv_maps_header_aliases_money_and_missing_regions() -> None:
    content = (
        "Ref,Sale Date,Customer,Item,Region,Amount\n"
        'R-001,02/06/2026,Gulberg Traders,Yarn,Punjab,"Rs. 45,900/-"\n'
        "R-002,10/06/2026,Al-Habib Stationers,Cloth,Sindh,12000\n"
        'R-003,20/07/2026,Indus Power,Yarn,,"(1,200)"\n'
    ).encode("utf-8")

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert [record.sales_row_id for record in records] == ["R-001", "R-002", "R-003"]
    # dayfirst: the Pakistani 02/06/2026 is 2 June, not 6 February.
    assert [record.date for record in records] == [
        date(2026, 6, 2),
        date(2026, 6, 10),
        date(2026, 7, 20),
    ]
    assert [record.amount for record in records] == [
        Decimal("45900"),
        Decimal("12000"),
        Decimal("-1200"),  # accounting-style parentheses are negative
    ]
    assert records[0].region == "Punjab"
    assert records[2].region is None  # an empty cell is no region, not ""
    # Provenance is the spreadsheet row, and no model was involved.
    assert records[0].source.document_id == "SLS-001"
    assert records[0].source.row_number == 2
    assert records[0].source.page is None
    # The report names the client's own headers behind each field.
    assert report.columns == {
        "sales_row_id": "Ref",
        "date": "Sale Date",
        "amount": "Amount",
        "customer_name": "Customer",
        "product": "Item",
        "region": "Region",
    }
    assert (report.format, report.encoding, report.delimiter) == ("csv", "utf-8-sig", ",")
    assert (report.header_row, report.rows_seen, report.rows_used, report.rows_skipped) == (1, 3, 3, 0)
    assert report.amount_derived is False


def test_reading_an_excel_export_parses_numeric_amounts() -> None:
    buffer = io.BytesIO()
    pd.DataFrame(
        {
            "Date": ["03/06/2026", "09/06/2026"],
            "Customer": ["Gulberg Traders", "Indus Power"],
            "Product": ["Yarn", "Cloth"],
            "Region": ["Punjab", None],
            "Amount": [45900, 120000],
        }
    ).to_excel(buffer, index=False)

    records, report = analytics.read_sales_export("SLS-009", "sales.xlsx", buffer.getvalue())

    assert [record.amount for record in records] == [Decimal("45900"), Decimal("120000")]
    assert records[1].region is None
    # No id column: rows are numbered as a human sees them, header is row 1.
    assert [record.sales_row_id for record in records] == ["SAL-0002", "SAL-0003"]
    assert report.format == "excel"
    assert report.sheet == "Sheet1"


def test_a_missing_amount_column_is_derived_from_quantity_and_unit_price() -> None:
    content = (
        "Date,Customer,Product,Qty,Rate\n"
        "02/06/2026,Gulberg Traders,Yarn,10,450.50\n"
        '03/06/2026,Indus Power,Cloth,3,"Rs. 1,000"\n'
    ).encode("utf-8")

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert [record.amount for record in records] == [Decimal("4505.00"), Decimal("3000")]
    assert report.amount_derived is True
    assert report.columns["quantity"] == "Qty"
    assert report.columns["unit_price"] == "Rate"


def test_no_amount_and_no_quantity_price_pair_is_rejected() -> None:
    content = b"Date,Customer,Product\n02/06/2026,Gulberg Traders,Yarn\n"

    with pytest.raises(analytics.SalesReadError, match="amount"):
        analytics.read_sales_export("SLS-001", "sales.csv", content)


def test_missing_customer_and_product_columns_are_filed_as_unspecified() -> None:
    """A bare date/amount export still counts toward revenue; the rows are not
    dropped, they are labelled, and the report says how many."""
    content = b"Date,Amount\n02/06/2026,100\n03/06/2026,250\n"

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert [record.customer_name for record in records] == [analytics.UNSPECIFIED] * 2
    assert [record.product for record in records] == [analytics.UNSPECIFIED] * 2
    assert report.filled_defaults == {"customer_name": 2, "product": 2}


def test_reading_sales_data_rejects_an_unsupported_format() -> None:
    with pytest.raises(analytics.SalesReadError, match="unsupported"):
        analytics.read_sales_export("SLS-001", "sales.pdf", b"%PDF-1.7")


def test_reading_sales_data_refuses_an_unquoted_comma_in_an_amount() -> None:
    """Left unquoted, pandas would read `Rs. 45` as the amount — or shift every
    column left. Silently auditing wrong numbers is the one thing this module
    must not do, so a ragged row is refused with instructions."""
    content = (
        "Sale Date,Customer,Item,Region,Amount\n"
        "02/06/2026,Gulberg Traders,Yarn,Punjab,Rs. 45,900/-\n"
    ).encode("utf-8")

    with pytest.raises(analytics.SalesReadError, match="quoted"):
        analytics.read_sales_export("SLS-001", "sales.csv", content)


def test_semicolon_and_tab_delimited_exports_are_read() -> None:
    semicolon = (
        "Date;Customer;Item;Amount\n"
        "02/06/2026;Gulberg Traders;Yarn;45900\n"
        "10/06/2026;Indus Power;Cloth;12000\n"
    ).encode("utf-8")
    tab = (
        "Date\tCustomer\tItem\tAmount\n"
        "02/06/2026\tGulberg Traders\tYarn\t45900\n"
    ).encode("utf-8")

    _semi_records, semi_report = analytics.read_sales_export("SLS-001", "sales.csv", semicolon)
    tab_records, tab_report = analytics.read_sales_export("SLS-002", "sales.tsv", tab)

    assert semi_report.delimiter == ";"
    assert semi_report.rows_used == 2
    assert tab_report.delimiter == "\t"
    assert tab_report.format == "tsv"
    assert tab_records[0].amount == Decimal("45900")


def test_title_rows_above_the_header_are_skipped_and_provenance_stays_true() -> None:
    content = (
        "Sales Report - Haroon Textiles\n"
        "\n"
        "Date,Customer,Item,Amount\n"
        "02/06/2026,Gulberg Traders,Yarn,100\n"
        "03/06/2026,Indus Power,Cloth,200\n"
    ).encode("utf-8")

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert report.header_row == 3
    # The rows are numbered as the spreadsheet shows them, title block included.
    assert [record.source.row_number for record in records] == [4, 5]
    assert [record.sales_row_id for record in records] == ["SAL-0004", "SAL-0005"]


def test_total_and_blank_rows_are_skipped_and_counted() -> None:
    content = (
        "Date,Customer,Item,Amount\n"
        "02/06/2026,Gulberg Traders,Yarn,100\n"
        "\n"
        "03/06/2026,Indus Power,Cloth,200\n"
        ",Total,,300\n"
        ",Grand Total,,300\n"
    ).encode("utf-8")

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert len(records) == 2
    assert report.skipped == {"blank": 1, "total_row": 2}
    assert report.rows_seen == 5 and report.rows_used == 2 and report.rows_skipped == 3


def test_a_windows_1252_csv_is_decoded() -> None:
    content = (
        "Date,Customer,Item,Amount\n02/06/2026,Café Traders,Yarn,100\n"
    ).encode("cp1252")

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert records[0].customer_name == "Café Traders"
    assert report.encoding == "cp1252"


def test_a_workbook_picks_the_sheet_that_holds_the_table() -> None:
    content = a_workbook(
        {
            "Notes": [["Prepared by accounts"], ["Do not edit"]],
            "Sales": [
                ["Invoice No", "Date", "Customer", "Product", "City", "Net Amount"],
                ["INV-1", datetime(2026, 6, 2), "Gulberg Traders", "Yarn", "Lahore", 45900],
                ["INV-2", datetime(2026, 6, 10), "Indus Power", "Cloth", "Karachi", 12000],
            ],
        }
    )

    records, report = analytics.read_sales_export("SLS-001", "sales.xlsx", content)

    assert report.sheet == "Sales"
    assert any("Notes" in warning for warning in report.warnings)
    assert [record.sales_row_id for record in records] == ["INV-1", "INV-2"]
    assert records[0].date == date(2026, 6, 2)
    assert records[0].region == "Lahore"
    assert report.columns["amount"] == "Net Amount"


def test_excel_serial_dates_that_lost_their_format_are_understood() -> None:
    content = a_workbook(
        {"Sheet1": [["Date", "Customer", "Item", "Amount"], [45810, "Gulberg Traders", "Yarn", 100]]}
    )

    records, _report = analytics.read_sales_export("SLS-001", "sales.xlsx", content)

    assert records[0].date == date(2025, 6, 2)


def test_a_header_that_only_hints_at_its_fields_is_still_found() -> None:
    """`Sale Date (DD/MM)`, `Customer / Party`, `Net Sales Amount` match no
    alias exactly; the hints resolve them — and `Total Tax` is left alone."""
    content = (
        "Sale Date (DD/MM),Customer / Party,Item Description,Total Tax,Net Sales Amount\n"
        "02/06/2026,Gulberg Traders,Yarn,900,45900\n"
    ).encode("utf-8")

    records, report = analytics.read_sales_export("SLS-001", "sales.csv", content)

    assert records[0].amount == Decimal("45900")
    assert report.columns["amount"] == "Net Sales Amount"
    assert report.columns["customer_name"] == "Customer / Party"
    assert report.columns["product"] == "Item Description"


def test_json_exports_are_read_as_a_list_or_under_a_data_key() -> None:
    rows = [
        {"date": "2026-06-02", "customer": "Gulberg Traders", "product": "Yarn", "amount": 45900},
        {"date": "2026-06-10", "customer": "Indus Power", "product": "Cloth", "amount": "12,000"},
    ]

    as_list, list_report = analytics.read_sales_export(
        "SLS-001", "sales.json", json.dumps(rows).encode("utf-8")
    )
    as_object, _ = analytics.read_sales_export(
        "SLS-002", "sales.json", json.dumps({"data": rows}).encode("utf-8")
    )

    assert [record.amount for record in as_list] == [Decimal("45900"), Decimal("12000")]
    assert [record.amount for record in as_object] == [Decimal("45900"), Decimal("12000")]
    assert list_report.format == "json"
    assert as_list[0].date == date(2026, 6, 2)


def test_an_opendocument_spreadsheet_is_read() -> None:
    buffer = io.BytesIO()
    pd.DataFrame(
        {
            "Date": ["02/06/2026", "10/06/2026"],
            "Customer": ["Gulberg Traders", "Indus Power"],
            "Product": ["Yarn", "Cloth"],
            "Amount": [45900, 12000],
        }
    ).to_excel(buffer, index=False, engine="odf")

    records, report = analytics.read_sales_export("SLS-001", "sales.ods", buffer.getvalue())

    assert report.format == "ods"
    assert [record.amount for record in records] == [Decimal("45900"), Decimal("12000")]


def test_a_file_with_no_recognisable_header_says_what_it_saw() -> None:
    content = b"alpha,beta,gamma\n1,2,3\n"

    with pytest.raises(analytics.SalesReadError, match="header row") as excinfo:
        analytics.read_sales_export("SLS-001", "sales.csv", content)
    assert "alpha, beta, gamma" in str(excinfo.value)


def test_read_sales_data_returns_the_records_alone() -> None:
    records = analytics.read_sales_data("SLS-001", "sales.csv", a_sales_csv())

    assert len(records) == 3
    assert sum((record.amount for record in records), Decimal(0)) == Decimal("87900")


# --------------------------------------------------------------------------- #
# The analysis
# --------------------------------------------------------------------------- #


def test_analyze_sales_partitions_the_records_by_month_and_product() -> None:
    records = [
        a_sale("SAL-0001", date(2026, 6, 2), "Gulberg Traders", "Yarn", "1000", "Punjab"),
        a_sale("SAL-0002", date(2026, 6, 10), "Al-Habib Stationers", "Yarn", "500", "Sindh"),
        a_sale("SAL-0003", date(2026, 6, 15), "Indus Power", "Cloth", "2000", "Punjab"),
        a_sale("SAL-0004", date(2026, 7, 1), "Gulberg Traders", "Cloth", "1500"),
        a_sale("SAL-0005", date(2026, 7, 20), "Indus Power", "Yarn", "300", "Sindh"),
    ]

    result = analytics.analyze_sales(records)

    assert result.record_count == 5
    assert result.total_revenue == Decimal("5300")
    assert result.period_start == date(2026, 6, 2)
    assert result.period_end == date(2026, 7, 20)
    assert result.document_ids == ["SLS-0001"]

    # Months partition the records exactly, and ascend.
    assert [(m.month, m.revenue, m.transaction_count) for m in result.monthly_revenue] == [
        ("2026-06", Decimal("3500"), 3),
        ("2026-07", Decimal("1800"), 2),
    ]

    # Products partition the records exactly, ranked by revenue.
    assert [(p.product, p.revenue, p.transaction_count) for p in result.revenue_by_product] == [
        ("Cloth", Decimal("3500"), 2),
        ("Yarn", Decimal("1800"), 3),
    ]
    assert [p.share for p in result.revenue_by_product] == [66.04, 33.96]


def test_analyze_sales_ranks_customers_and_regions() -> None:
    records = [
        a_sale("SAL-0001", date(2026, 6, 2), "Gulberg Traders", "Yarn", "1000", "Punjab"),
        a_sale("SAL-0002", date(2026, 6, 10), "Al-Habib Stationers", "Yarn", "500", "Sindh"),
        a_sale("SAL-0003", date(2026, 6, 15), "Indus Power", "Cloth", "2000", "Punjab"),
        a_sale("SAL-0004", date(2026, 7, 1), "Gulberg Traders", "Cloth", "1500"),
        a_sale("SAL-0005", date(2026, 7, 20), "Indus Power", "Yarn", "300", "Sindh"),
    ]

    result = analytics.analyze_sales(records)

    assert [
        (c.customer_name, c.revenue, c.transaction_count) for c in result.top_customers
    ] == [
        ("Gulberg Traders", Decimal("2500"), 2),
        ("Indus Power", Decimal("2300"), 2),
        ("Al-Habib Stationers", Decimal("500"), 1),
    ]
    # Only records that carry a region are counted — SAL-0004 has none.
    assert [(r.region, r.revenue, r.transaction_count) for r in result.sales_by_region] == [
        ("Punjab", Decimal("3000"), 2),
        ("Sindh", Decimal("800"), 2),
    ]
    # A clean month with clean rows raises nothing.
    assert result.anomalies == []


def test_top_customers_is_capped_at_five() -> None:
    records = [
        a_sale(
            f"SAL-{position:04d}",
            date(2026, 6, position),
            f"Customer {position}",
            "Yarn",
            "100",
        )
        for position in range(1, 7)  # six customers, equal revenue
    ]

    result = analytics.analyze_sales(records)

    assert len(result.top_customers) == 5
    # Equal revenue: the tie breaks alphabetically, deterministically.
    assert [c.customer_name for c in result.top_customers] == [
        "Customer 1",
        "Customer 2",
        "Customer 3",
        "Customer 4",
        "Customer 5",
    ]


def test_analyze_sales_flags_duplicate_and_negative_rows() -> None:
    records = [
        a_sale("SAL-0001", date(2026, 6, 2), "Gulberg Traders", "Yarn", "1000", "Punjab"),
        a_sale("SAL-0002", date(2026, 6, 2), "Gulberg Traders", "Yarn", "1000", "Punjab"),
        a_sale("SAL-0003", date(2026, 6, 5), "Indus Power", "Cloth", "-200", "Sindh", row_number=4),
    ]

    result = analytics.analyze_sales(records)

    by_kind = {anomaly.kind: anomaly for anomaly in result.anomalies}
    assert set(by_kind) == {"negative-amount", "duplicate-transaction"}

    duplicate = by_kind["duplicate-transaction"]
    assert duplicate.source_row_id == "SAL-0001"
    assert duplicate.related_row_ids == ["SAL-0001", "SAL-0002"]

    negative = by_kind["negative-amount"]
    assert negative.source_row_id == "SAL-0003"
    assert negative.source is not None
    assert negative.source.row_number == 4


def test_analyze_sales_flags_a_month_far_from_the_median() -> None:
    records = [
        a_sale("SAL-0001", date(2026, 4, 10), "April Buyer", "Yarn", "100"),
        a_sale("SAL-0002", date(2026, 5, 10), "May Buyer", "Yarn", "100"),
        a_sale("SAL-0003", date(2026, 6, 10), "June Buyer", "Yarn", "500"),
    ]

    result = analytics.analyze_sales(records)

    assert [anomaly.kind for anomaly in result.anomalies] == ["revenue-spike"]
    assert result.anomalies[0].month == "2026-06"
    assert result.anomalies[0].source_row_id is None


def test_analyze_sales_flags_a_transaction_far_above_the_median() -> None:
    records = [
        a_sale(f"SAL-{position:04d}", date(2026, 6, 10), f"Buyer {position}", "Yarn", "100")
        for position in range(1, 10)
    ] + [a_sale("SAL-0010", date(2026, 6, 11), "Wholesale Mart", "Yarn", "2000")]

    result = analytics.analyze_sales(records)

    assert [anomaly.kind for anomaly in result.anomalies] == ["large-transaction"]
    assert result.anomalies[0].source_row_id == "SAL-0010"
    # The nine ordinary sales are not flagged — the rule needs its sample first.
    assert result.record_count == 10


def test_analyze_sales_on_empty_input_is_an_empty_readout() -> None:
    result = analytics.analyze_sales([])

    assert result.record_count == 0
    assert result.total_revenue == Decimal("0")
    assert result.monthly_revenue == []
    assert result.anomalies == []
    assert result.period_start is None


def test_the_readout_carries_the_data_quality_reports_it_was_read_from() -> None:
    records, report = analytics.read_sales_export("SLS-001", "sales.csv", a_sales_csv())

    result = analytics.analyze_sales(records, reports=[report])

    assert result.data_quality == [report]
    assert result.data_quality[0].rows_used == 3


def test_the_result_survives_a_round_trip_through_the_store(
    repository: SqliteCaseRepository, demo_org: str
) -> None:
    records, report = analytics.read_sales_export("SLS-0001", "sales.csv", a_sales_csv())
    records.append(a_sale("SAL-9999", date(2026, 7, 1), "Gulberg Traders", "Cloth", "-50"))
    create_case(repository, demo_org, "CASE-SLS-001")
    result = analytics.analyze_sales(records, reports=[report])

    repository.save_sales_analytics(demo_org, "CASE-SLS-001", result)
    restored = repository.get_sales_analytics(demo_org, "CASE-SLS-001")

    assert restored == result
    assert restored is not None
    assert restored.total_revenue == Decimal("87850")
    assert restored.data_quality[0].columns["amount"] == "Amount"
    # Another firm's lookup finds nothing, not somebody else's readout.
    assert repository.get_sales_analytics("11111111-1111-4111-8111-111111111111", "CASE-SLS-001") is None
    # A re-run replaces, which is what upsert-on-(org, case) is for.
    repository.save_sales_analytics(demo_org, "CASE-SLS-001", analytics.analyze_sales([]))
    assert repository.get_sales_analytics(demo_org, "CASE-SLS-001").record_count == 0


# --------------------------------------------------------------------------- #
# Exporting the readout
# --------------------------------------------------------------------------- #


def test_export_workbook_copies_every_breakdown_onto_its_own_sheet() -> None:
    records, report = analytics.read_sales_export("SLS-001", "sales.csv", a_sales_csv())
    result = analytics.analyze_sales(records, reports=[report])

    workbook = load_workbook(io.BytesIO(analytics.export_workbook(result)))

    assert workbook.sheetnames == [
        "Summary", "Monthly revenue", "By product", "By region", "Top customers",
        "Anomalies", "Data quality",
    ]
    monthly = list(workbook["Monthly revenue"].iter_rows(values_only=True))
    assert monthly[0] == ("Month", "Revenue", "Transactions")
    assert monthly[1:] == [("2026-06", 57900, 2), ("2026-07", 30000, 1)]
    summary = dict(workbook["Summary"].iter_rows(values_only=True))
    assert summary["Total revenue"] == 87900
    assert summary["Sales records"] == 3
    quality = list(workbook["Data quality"].iter_rows(values_only=True))
    assert quality[1][0] == "sales.csv"
    assert quality[1][6] == 3  # rows used


# --------------------------------------------------------------------------- #
# The routes
# --------------------------------------------------------------------------- #


def test_uploading_a_sales_export_stores_it_and_lists_it(
    client: TestClient, repository: SqliteCaseRepository, demo_org: str, seeded_case: str
) -> None:
    response = upload_sales(client, seeded_case, "sales.csv", a_sales_csv())

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["filename"] == "sales.csv"
    assert body["case_id"] == seeded_case
    assert body["uploaded_by"]

    listed = client.get(f"/v1/cases/{seeded_case}/sales-data")
    assert listed.status_code == 200
    assert [upload["sales_data_id"] for upload in listed.json()["uploads"]] == [body["sales_data_id"]]
    assert repository.list_sales_data_uploads(demo_org, seeded_case)[0].filename == "sales.csv"


def test_an_unreadable_export_is_refused_at_upload_with_the_reason(
    client: TestClient, seeded_case: str
) -> None:
    """A file the reader cannot use never reaches storage: the answer comes now,
    with the reason, not on a later run."""
    response = upload_sales(client, seeded_case, "sales.csv", b"alpha,beta\n1,2\n")

    assert response.status_code == 422, response.text
    assert "could not be read" in response.json()["detail"]
    assert "header row" in response.json()["detail"]
    assert client.get(f"/v1/cases/{seeded_case}/sales-data").json()["uploads"] == []


def test_an_unsupported_suffix_is_refused_before_it_is_read(
    client: TestClient, seeded_case: str
) -> None:
    response = upload_sales(client, seeded_case, "sales.pdf", b"%PDF-1.7")

    assert response.status_code == 415
    assert ".xlsx" in response.json()["detail"]


def test_post_runs_the_analysis_persists_it_and_records_the_trail(
    client, repository: SqliteCaseRepository, demo_org: str, storage: LocalDocumentStore
) -> None:
    case_id = seed_sales_case(repository, demo_org, storage)

    response = client.post(f"/v1/cases/{case_id}/analytics")

    assert response.status_code == 201, response.text
    body = response.json()
    assert body["record_count"] == 3
    assert Decimal(str(body["total_revenue"])) == Decimal("87900")
    assert [entry["month"] for entry in body["monthly_revenue"]] == ["2026-06", "2026-07"]
    assert body["data_quality"][0]["filename"] == "sales.csv"
    assert body["data_quality"][0]["rows_used"] == 3

    saved = repository.get_sales_analytics(demo_org, case_id)
    assert saved is not None
    assert saved.record_count == 3
    assert saved.document_ids == ["SLS-0001"]

    runs = [
        record
        for record in repository.list_audit(demo_org, case_id)
        if record.action is AuditAction.SALES_ANALYTICS_RUN
    ]
    assert len(runs) == 1
    assert runs[0].actor_type is ActorType.HUMAN
    assert runs[0].actor_id == USER_ID


def test_several_exports_are_read_in_upload_order_and_summed_whole(
    client, repository: SqliteCaseRepository, demo_org: str, storage: LocalDocumentStore
) -> None:
    case_id = seed_sales_case(repository, demo_org, storage)
    august = b"Date,Customer,Item,Amount\n05/08/2026,Gulberg Traders,Yarn,1000\n"
    seed_sales_upload(
        repository, demo_org, storage, case_id, sales_data_id="SLS-0002",
        filename="august.csv", content=august,
    )

    body = client.post(f"/v1/cases/{case_id}/analytics").json()

    assert body["record_count"] == 4
    assert Decimal(str(body["total_revenue"])) == Decimal("88900")
    assert [entry["month"] for entry in body["monthly_revenue"]] == ["2026-06", "2026-07", "2026-08"]
    assert [report["filename"] for report in body["data_quality"]] == ["sales.csv", "august.csv"]


def test_get_returns_the_saved_result_and_404s_before_any_run(
    client, repository: SqliteCaseRepository, demo_org: str, storage: LocalDocumentStore
) -> None:
    case_id = seed_sales_case(repository, demo_org, storage)

    missing = client.get(f"/v1/cases/{case_id}/analytics")
    assert missing.status_code == 404

    assert client.post(f"/v1/cases/{case_id}/analytics").status_code == 201
    cached = client.get(f"/v1/cases/{case_id}/analytics")

    assert cached.status_code == 200
    assert Decimal(str(cached.json()["total_revenue"])) == Decimal("87900")
    assert cached.json()["record_count"] == 3


def test_post_without_any_sales_data_is_rejected(client, seeded_case: str) -> None:
    """The Haroon sample case has no sales export: 422 with a plain sentence, not 500."""
    response = client.post(f"/v1/cases/{seeded_case}/analytics")

    assert response.status_code == 422
    assert "no sales data" in response.json()["detail"]


def test_deleting_an_export_removes_it_from_the_next_run(
    client, repository: SqliteCaseRepository, demo_org: str, storage: LocalDocumentStore
) -> None:
    case_id = seed_sales_case(repository, demo_org, storage)
    assert client.post(f"/v1/cases/{case_id}/analytics").json()["record_count"] == 3

    deleted = client.delete(f"/v1/cases/{case_id}/sales-data/SLS-0001")
    assert deleted.status_code == 204
    assert client.get(f"/v1/cases/{case_id}/sales-data").json()["uploads"] == []
    # With nothing left to read, a re-run is refused rather than saving zeros.
    assert client.post(f"/v1/cases/{case_id}/analytics").status_code == 422
    # A second delete of the same id is a 404, not a silent success.
    assert client.delete(f"/v1/cases/{case_id}/sales-data/SLS-0001").status_code == 404


def test_download_hands_back_the_readout_as_a_workbook_or_json(
    client, repository: SqliteCaseRepository, demo_org: str, storage: LocalDocumentStore
) -> None:
    case_id = seed_sales_case(repository, demo_org, storage)

    nothing_yet = client.get(f"/v1/cases/{case_id}/analytics/download")
    assert nothing_yet.status_code == 404

    assert client.post(f"/v1/cases/{case_id}/analytics").status_code == 201

    excel = client.get(f"/v1/cases/{case_id}/analytics/download")
    assert excel.status_code == 200, excel.text
    assert excel.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert f"{case_id}-sales-analytics.xlsx" in excel.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(excel.content))
    assert "Monthly revenue" in workbook.sheetnames
    assert dict(workbook["Summary"].iter_rows(values_only=True))["Total revenue"] == 87900

    as_json = client.get(f"/v1/cases/{case_id}/analytics/download", params={"format": "json"})
    assert as_json.status_code == 200
    assert as_json.headers["content-type"].startswith("application/json")
    assert as_json.json()["record_count"] == 3

    assert client.get(
        f"/v1/cases/{case_id}/analytics/download", params={"format": "pdf"}
    ).status_code == 422


def test_analytics_is_scoped_to_the_callers_organization(
    client,
    other_client,
    repository: SqliteCaseRepository,
    demo_org: str,
    storage: LocalDocumentStore,
) -> None:
    case_id = seed_sales_case(repository, demo_org, storage)
    assert client.post(f"/v1/cases/{case_id}/analytics").status_code == 201

    # Firm B: same store, same route, another tenant. The case is not theirs,
    # so it does not exist as far as they can tell — 404, whatever the verb.
    assert other_client.get(f"/v1/cases/{case_id}/analytics").status_code == 404
    assert other_client.post(f"/v1/cases/{case_id}/analytics").status_code == 404
    assert other_client.get(f"/v1/cases/{case_id}/sales-data").status_code == 404
    assert other_client.get(f"/v1/cases/{case_id}/analytics/download").status_code == 404


def test_upload_then_run_end_to_end_over_http(
    client: TestClient, repository: SqliteCaseRepository, demo_org: str, seeded_case: str
) -> None:
    """The path the Analytics screen takes: upload an export, run, read back."""
    content = a_workbook(
        {
            "Sales": [
                ["Date", "Customer", "Product", "Qty", "Unit Price"],
                [datetime(2026, 6, 2), "Gulberg Traders", "Yarn", 10, 100],
                [datetime(2026, 6, 3), "Indus Power", "Cloth", 2, 250.5],
            ]
        }
    )
    assert upload_sales(client, seeded_case, "june.xlsx", content).status_code == 201

    run = client.post(f"/v1/cases/{seeded_case}/analytics")
    assert run.status_code == 201, run.text
    body = run.json()
    assert body["record_count"] == 2
    assert Decimal(str(body["total_revenue"])) == Decimal("1501.0")
    assert body["data_quality"][0]["amount_derived"] is True

    assert client.get(f"/v1/cases/{seeded_case}/analytics").json()["record_count"] == 2
    assert repository.get_sales_analytics(demo_org, seeded_case) is not None


# --------------------------------------------------------------------------- #
# Dashboard and pipeline
# --------------------------------------------------------------------------- #


def test_dashboard_includes_sales_analytics_when_available(
    client: TestClient,
    repository: SqliteCaseRepository,
    demo_org: str,
    seeded_case: str,
) -> None:
    """When sales analytics have been saved, the dashboard returns them."""
    records = [
        a_sale("SAL-0001", date(2026, 6, 2), "Gulberg Traders", "Yarn", "1000", "Punjab"),
        a_sale("SAL-0002", date(2026, 6, 10), "Al-Habib Stationers", "Yarn", "500", "Sindh"),
        a_sale("SAL-0003", date(2026, 7, 1), "Indus Power", "Cloth", "1500"),
    ]
    result = analytics.analyze_sales(records)
    repository.save_sales_analytics(demo_org, seeded_case, result)

    response = client.get("/v1/dashboard")
    assert response.status_code == 200
    body = response.json()

    assert body["sales_analytics"] is not None
    assert body["sales_analytics"]["record_count"] == 3
    assert Decimal(str(body["sales_analytics"]["total_revenue"])) == Decimal("3000")


def test_dashboard_returns_null_sales_analytics_when_none_saved(
    client: TestClient, seeded_case: str
) -> None:
    """A case without sales data has null sales_analytics on the dashboard."""
    response = client.get("/v1/dashboard")
    assert response.status_code == 200
    assert response.json()["sales_analytics"] is None


def test_the_case_pipeline_never_runs_sales_analytics(
    repository: SqliteCaseRepository, storage: LocalDocumentStore, demo_mode
) -> None:
    """Sales exports are a separate data source: the audit pipeline over a
    ledger produces a review queue and nothing under analytics."""
    case_id = "CASE-NOSALES"
    documents = [
        (
            StoredDocument(
                document_id="DOC-LED-001",
                document_type=DocumentType.LEDGER,
                filename="ledger.xlsx",
                size_bytes=1,
                storage_path=f"{case_id}/DOC-LED-001/ledger.xlsx",
            ),
            a_ledger(),
        ),
    ]

    outcome = run_pipeline(
        ORG, case_id, "Client", documents, AUDITOR, repository, storage
    )

    assert outcome.status is CaseStatus.READY_FOR_REVIEW
    assert repository.get_sales_analytics(ORG, case_id) is None
    assert all(
        record.action is not AuditAction.SALES_ANALYTICS_RUN
        for record in repository.list_audit(ORG, case_id)
    )


# --------------------------------------------------------------------------- #
# The module contract
# --------------------------------------------------------------------------- #


def test_the_analytics_module_imports_no_ai_client_and_no_other_module() -> None:
    """Deterministic and self-contained, checked by hand the way `test_rules.py`
    checks `rules/`: no AI client anywhere in the package, and no other module
    either — the package stands on `app.shared/` alone."""
    forbidden = {"httpx", "openai", "dashscope", "anthropic", "requests"}
    package = Path(analytics.__file__).parent
    for path in package.glob("*.py"):
        tree = ast.parse(path.read_text("utf-8"))
        for node in ast.walk(tree):
            names: list[str] = []
            if isinstance(node, ast.Import):
                names = [alias.name for alias in node.names]
            elif isinstance(node, ast.ImportFrom) and node.module:
                names = [node.module]
            for name in names:
                assert name.split(".")[0] not in forbidden, f"{path.name} imports {name}"
                assert not name.startswith("app.modules"), f"{path.name} imports {name}"
