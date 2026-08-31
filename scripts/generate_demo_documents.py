"""Generate the demo documents for the Haroon Textiles sample case.

Creates `sample-data/demo-documents/` with everything an upload needs to run
the documented sample case through the real pipeline (upload -> extract ->
match -> flag -> review):

- `haroon-textiles-ledger-june-2026.xlsx` - the client ledger, read by pandas.
- `almadina-bank-statement-june-2026.pdf` - a two-page statement with the nine
  payments the ledger should reconcile against.
- `invoice-karachi-packaging-INV-2026-0087.pdf` - paid twice on purpose.
- `invoice-sialkot-metal-works-SMW-2026-0431.pdf` - the loom spare parts invoice.

The documents carry the sample case's five planted errors:

1. Digit transposition: the ledger books 45,900 to Al-Habib Stationers while
   the cheque cleared 49,500 on the same day.
2. Duplicate invoice payment: INV-2026-0087 (Karachi Packaging Co.) is paid on
   05-06 and again on 16-06.
3. Structuring / near-limit: two payments of 49,500 to Hussain Brothers & Sons
   on the same day, each just under the 50,000 approval limit.
4. Weekend round-number posting: 1,500,000 to Indus Power Solutions on Sunday
   14-06-2026.
5. Fictitious vendor: 187,500 to Shalimar Trading Co with no bank payment and
   no invoice anywhere in the documents.

Run from the repo root:

    backend/.venv/Scripts/python scripts/generate_demo_documents.py    (Windows)
    backend/.venv/bin/python scripts/generate_demo_documents.py        (Linux/Mac)

Add `--check` to re-read the generated files with the app's own readers and
confirm they parse. Output is deterministic.
"""

from __future__ import annotations

import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = REPO_ROOT / "sample-data" / "demo-documents"

GREEN = "14532d"

LEDGER_FILE = "haroon-textiles-ledger-june-2026.xlsx"
STATEMENT_FILE = "almadina-bank-statement-haroon-textiles-june-2026.pdf"
INVOICE_KP_FILE = "invoice-karachi-packaging-INV-2026-0087.pdf"
INVOICE_SMW_FILE = "invoice-sialkot-metal-works-SMW-2026-0431.pdf"

#: (date, voucher no, party, description, account code, amount in PKR)
LEDGER_ROWS = [
    (date(2026, 6, 2),  "LED-0003", "Gulberg Traders (Pvt) Ltd", "Yarn purchase - June lot 1",          "5010", 284_000),
    (date(2026, 6, 5),  "LED-0007", "Karachi Packaging Co.",     "Carton supply against INV-2026-0087", "5020",  96_400),
    (date(2026, 6, 8),  "LED-0009", "Ravi Logistics Pvt Ltd",    "Freight - Lahore to Karachi",         "5310",  63_750),
    (date(2026, 6, 10), "LED-0012", "Al-Habib Stationers",       "Office supplies - Q2",                "6110",  45_900),
    (date(2026, 6, 11), "LED-0014", "Hussain Brothers & Sons",   "Dyeing services - part 1",            "5040",  49_500),
    (date(2026, 6, 11), "LED-0015", "Hussain Brothers & Sons",   "Dyeing services - part 2",            "5040",  49_500),
    (date(2026, 6, 14), "LED-0019", "Indus Power Solutions",     "Generator overhaul - advance",        "1720", 1_500_000),
    (date(2026, 6, 16), "LED-0023", "Karachi Packaging Co.",     "Carton supply against INV-2026-0087", "5020",  96_400),
    (date(2026, 6, 17), "LED-0027", "Sialkot Metal Works",       "Loom spare parts",                    "5050", 312_880),
    (date(2026, 6, 18), "LED-0031", "Shalimar Trading Co",       "Consultancy - brand audit",           "6420", 187_500),
]

#: (date, description, amount in PKR). Balances are derived from the opening
#: balance so the running balance column actually adds up.
STATEMENT_ROWS = [
    (date(2026, 6, 2),  "IBFT GULBERG TRADERS PVT LTD",     284_000),
    (date(2026, 6, 5),  "CHQ 004412 KARACHI PACKAGING CO",   96_400),
    (date(2026, 6, 10), "ONLINE TFR RAVI LOGISTICS",         63_750),
    (date(2026, 6, 10), "CHQ 004418 AL HABIB STATIONERS",    49_500),
    (date(2026, 6, 11), "IBFT HUSSAIN BROTHERS AND SONS",    49_500),
    (date(2026, 6, 11), "IBFT HUSSAIN BROTHERS AND SONS",    49_500),
    (date(2026, 6, 14), "RTGS INDUS POWER SOLUTIONS",     1_500_000),
    (date(2026, 6, 16), "CHQ 004431 KARACHI PACKAGING CO",   96_400),
    (date(2026, 6, 19), "IBFT SIALKOT METAL WORKS",         312_880),
]

OPENING_BALANCE = Decimal("5105330")

BANK_NAME = "AL-MADINA BANK LIMITED"
ACCOUNT_TITLE = "HAROON TEXTILES (PVT) LTD"
ACCOUNT_NO = "0201-0109482730"
IBAN = "PK36ALMD0002010109482730"


def _money(value: int | Decimal) -> str:
    return f"{value:,.2f}"


def _dmy(when: date) -> str:
    return when.strftime("%d-%m-%Y")


def _statement_balances() -> list[tuple[date, str, int, Decimal]]:
    balance = OPENING_BALANCE
    rows: list[tuple[date, str, int, Decimal]] = []
    for when, description, amount in STATEMENT_ROWS:
        balance -= Decimal(amount)
        rows.append((when, description, amount, balance))
    return rows


# --------------------------------------------------------------------------- #
# The ledger
# --------------------------------------------------------------------------- #


def _write_ledger(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "June 2026"

    headers = ["Date", "Voucher No", "Party Name", "Description", "Account Code", "Amount (PKR)"]
    header_fill = PatternFill("solid", fgColor=GREEN)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, (when, voucher, party, description, code, amount) in enumerate(LEDGER_ROWS, start=2):
        sheet.cell(row=row, column=1, value=when).number_format = "DD-MM-YYYY"
        sheet.cell(row=row, column=2, value=voucher)
        sheet.cell(row=row, column=3, value=party)
        sheet.cell(row=row, column=4, value=description)
        sheet.cell(row=row, column=5, value=code)
        sheet.cell(row=row, column=6, value=amount).number_format = "#,##0.00"

    for column, width in zip("ABCDEF", (12, 13, 30, 40, 14, 14)):
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"

    workbook.save(path)


# --------------------------------------------------------------------------- #
# The bank statement
# --------------------------------------------------------------------------- #


def _statement_table(rows: list[tuple[date, str, int, Decimal]]):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    data = [["DATE", "DESCRIPTION", "AMOUNT (PKR)", "BALANCE (PKR)"]]
    for when, description, amount, balance in rows:
        data.append([_dmy(when), description, _money(amount), _money(balance)])

    table = Table(data, colWidths=[62, 235, 100, 110], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{GREEN}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b7c4c9")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f6f7")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return table


def _statement_page(canvas, doc) -> None:
    """The letterhead band, account details, and footer drawn on every page."""
    from reportlab.lib import colors

    width, height = canvas._pagesize
    canvas.saveState()

    canvas.setFillColor(colors.HexColor(f"#{GREEN}"))
    canvas.rect(0, height - 74, width, 74, stroke=0, fill=1)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 16)
    canvas.drawString(40, height - 34, BANK_NAME)
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(width - 40, height - 34, "ACCOUNT STATEMENT")

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 9.5)
    canvas.drawString(40, height - 92, f"Account Title: {ACCOUNT_TITLE}")
    canvas.setFont("Helvetica", 9)
    canvas.drawString(40, height - 106, f"Account No: {ACCOUNT_NO}    IBAN: {IBAN}    Currency: PKR")
    canvas.drawString(40, height - 120, "Statement period: 01-06-2026 to 30-06-2026")

    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawCentredString(width / 2, 26, f"Page {doc.page} of 2 - {BANK_NAME}")
    canvas.restoreState()


def _write_statement(path: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        PageBreak,
        PageTemplate,
        Paragraph,
        Spacer,
    )

    rows = _statement_balances()
    body = ParagraphStyle("body", fontName="Helvetica", fontSize=9.5)
    small = ParagraphStyle(
        "small", fontName="Helvetica", fontSize=8.5, textColor=colors.HexColor("#555555")
    )

    doc = BaseDocTemplate(
        str(path), pagesize=A4, title="Account Statement - June 2026", author=ACCOUNT_TITLE
    )
    frame = Frame(
        40, 55, A4[0] - 80, A4[1] - 190, leftPadding=0, rightPadding=0,
        topPadding=0, bottomPadding=0,
    )
    doc.addPageTemplates([PageTemplate(id="statement", frames=[frame], onPage=_statement_page)])

    story = [
        Paragraph(f"Opening Balance as of 01-06-2026: <b>PKR {_money(OPENING_BALANCE)}</b>", body),
        Spacer(1, 12),
        _statement_table(rows[:5]),
        PageBreak(),
        _statement_table(rows[5:]),
        Spacer(1, 14),
        Paragraph(f"Closing Balance as of 30-06-2026: <b>PKR {_money(rows[-1][3])}</b>", body),
        Spacer(1, 8),
        Paragraph(
            "Please examine this statement and report any discrepancy within 14 days. "
            "This is a computer-generated statement and does not require a signature.",
            small,
        ),
    ]
    doc.build(story)


# --------------------------------------------------------------------------- #
# The invoices
# --------------------------------------------------------------------------- #


def _write_invoice(
    path: Path,
    *,
    vendor: str,
    address: str,
    ntn: str,
    number: str,
    when: str,
    items: list[tuple[str, int, float, int]],
    words: str,
) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdfcanvas

    total = sum(amount for _description, _qty, _rate, amount in items)

    canvas = pdfcanvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    left, right = 40, width - 40

    # Letterhead band.
    canvas.setFillColor(colors.HexColor(f"#{GREEN}"))
    canvas.rect(0, height - 74, width, 74, stroke=0, fill=1)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 18)
    canvas.drawString(left, height - 34, vendor)
    canvas.setFont("Helvetica", 9)
    canvas.drawString(left, height - 52, address)
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawRightString(right, height - 34, "SALES INVOICE")

    # Meta block.
    canvas.setFillColor(colors.black)
    y = height - 100
    canvas.setFont("Helvetica-Bold", 10.5)
    canvas.drawString(left, y, f"Invoice No.: {number}")
    canvas.drawRightString(right, y, f"Date: {when}")
    y -= 15
    canvas.setFont("Helvetica", 9.5)
    canvas.drawString(left, y, "Bill To: Haroon Textiles (Pvt) Ltd, 12-A Gulberg III, Lahore")

    # Items table.
    y -= 34
    col_qty, col_rate, col_amount = left + 320, left + 375, right
    canvas.setFillColor(colors.HexColor("#e8eef0"))
    canvas.rect(left, y - 5, right - left, 18, stroke=0, fill=1)
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(left + 6, y, "Description")
    canvas.drawRightString(col_qty, y, "Qty")
    canvas.drawRightString(col_rate, y, "Rate (PKR)")
    canvas.drawRightString(col_amount, y, "Amount (PKR)")

    y -= 22
    canvas.setFont("Helvetica", 9)
    for description, qty, rate, amount in items:
        canvas.drawString(left + 6, y, description)
        canvas.drawRightString(col_qty, y, f"{qty:,}")
        canvas.drawRightString(col_rate, y, f"{rate:,.2f}")
        canvas.drawRightString(col_amount, y, f"{amount:,.2f}")
        canvas.setStrokeColor(colors.HexColor("#d7dee0"))
        canvas.setLineWidth(0.4)
        canvas.line(left, y - 6, right, y - 6)
        y -= 22

    # Totals.
    y -= 16
    canvas.setFont("Helvetica", 10)
    canvas.drawRightString(right, y, f"Subtotal: {total:,.2f}")
    y -= 15
    canvas.drawRightString(right, y, "Sales tax: included in total")
    y -= 24
    canvas.setFillColor(colors.HexColor(f"#{GREEN}"))
    canvas.rect(right - 250, y - 8, 250, 28, stroke=0, fill=1)
    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawRightString(right - 10, y, f"Total: Rs. {total:,.0f}/-")

    y -= 32
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Oblique", 9)
    canvas.drawString(left, y, f"Amount in words: {words}")

    # Footer.
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawString(left, 60, f"NTN: {ntn}")
    canvas.drawString(left, 48, f"Payment due within 30 days. Cheques payable to {vendor}.")
    canvas.drawCentredString(width / 2, 34, "Computer-generated invoice - no signature required.")

    canvas.showPage()
    canvas.save()


# --------------------------------------------------------------------------- #
# Check: re-read the files with the app's own readers
# --------------------------------------------------------------------------- #


def _check(ledger: Path, statement: Path, invoice_kp: Path, invoice_smw: Path) -> None:
    sys.path.insert(0, str(REPO_ROOT / "backend"))
    from app.modules.extraction.ledger_reader import read_ledger
    from app.modules.extraction.page_images import pdf_page_count

    entries = read_ledger("DOC-LED-001", ledger.name, ledger.read_bytes())
    assert len(entries) == len(LEDGER_ROWS), f"expected {len(LEDGER_ROWS)} rows, got {len(entries)}"
    for entry, expected in zip(entries, LEDGER_ROWS):
        when, voucher, party, _description, _code, amount = expected
        assert entry.ledger_row_id == voucher, f"{entry.ledger_row_id} != {voucher}"
        assert entry.date == when, f"{entry.date} != {when}"
        assert entry.amount == Decimal(amount), f"{entry.amount} != {amount}"
        assert entry.party_name == party
    print(f"  ledger: {len(entries)} rows read by pandas, ids and amounts as planted")

    assert pdf_page_count(statement.read_bytes()) == 2, "statement should have 2 pages"
    print("  statement: 2 pages, opens as a PDF")

    for path, expected_pages in ((invoice_kp, 1), (invoice_smw, 1)):
        assert pdf_page_count(path.read_bytes()) == expected_pages
        print(f"  {path.name}: 1 page, opens as a PDF")


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #


def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    check = "--check" in argv

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    ledger = OUT_DIR / LEDGER_FILE
    statement = OUT_DIR / STATEMENT_FILE
    invoice_kp = OUT_DIR / INVOICE_KP_FILE
    invoice_smw = OUT_DIR / INVOICE_SMW_FILE

    _write_ledger(ledger)
    _write_statement(statement)
    _write_invoice(
        invoice_kp,
        vendor="KARACHI PACKAGING CO.",
        address="Plot 44, Sector 23, Korangi Industrial Area, Karachi",
        ntn="4271817-7",
        number="INV-2026-0087",
        when="03-06-2026",
        items=[("Corrugated cartons, 5-ply, 18x12x12 in. (June lot)", 1_600, 60.25, 96_400)],
        words="Rupees Ninety-Six Thousand Four Hundred Only",
    )
    _write_invoice(
        invoice_smw,
        vendor="SIALKOT METAL WORKS",
        address="Small Industrial Estate, Ugoki Road, Sialkot",
        ntn="3552209-4",
        number="SMW/2026/0431",
        when="15-06-2026",
        items=[
            ("Rapier tape assembly, 92 in.",     4, 28_500.00, 114_000),
            ("Gripper set, SMIT",                6, 14_500.00,  87_000),
            ("Cutter blade pack",               10,  2_988.00,  29_880),
            ("Spare pin kit",                    2, 41_000.00,  82_000),
        ],
        words="Rupees Three Hundred Twelve Thousand Eight Hundred Eighty Only",
    )

    for path in (ledger, statement, invoice_kp, invoice_smw):
        print(f"wrote {path.relative_to(REPO_ROOT)} ({path.stat().st_size:,} bytes)")

    if check:
        print()
        _check(ledger, statement, invoice_kp, invoice_smw)
        print()
        print("All demo documents parse. Upload them on the Documents page:")
        print(f"  bank statement = {statement.name}")
        print(f"  ledger         = {ledger.name}")
        print(f"  invoices       = {invoice_kp.name}, {invoice_smw.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
